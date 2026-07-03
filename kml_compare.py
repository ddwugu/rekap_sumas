import streamlit as st
import pandas as pd
import zipfile
import io
import os
from xml.etree import ElementTree as ET
from math import radians, cos, sin, asin, sqrt
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

st.set_page_config(page_title="KML/KMZ Overlap Checker", page_icon="🗺️", layout="wide")
KML_NS = "http://www.opengis.net/kml/2.2"


# ── Helpers ──────────────────────────────────────────────────────────────────

def filename_to_label(uploaded_file):
    if uploaded_file is None:
        return ""
    base = os.path.splitext(uploaded_file.name)[0]
    return base.replace("_", " ").replace("-", " ")


def haversine_m(lat1, lon1, lat2, lon2):
    R = 6_371_000
    phi1, phi2 = radians(lat1), radians(lat2)
    a = sin((phi2 - phi1) / 2) ** 2 + cos(phi1) * cos(phi2) * sin((radians(lon2 - lon1)) / 2) ** 2
    return 2 * R * asin(sqrt(a))


def extract_kml_bytes(uploaded_file):
    name = uploaded_file.name.lower()
    raw = uploaded_file.read()
    if name.endswith(".kmz"):
        with zipfile.ZipFile(io.BytesIO(raw)) as z:
            kml_names = [n for n in z.namelist() if n.lower().endswith(".kml")]
            if not kml_names:
                st.error(f"Tidak ada file .kml di dalam {uploaded_file.name}")
                return None
            return z.read(kml_names[0])
    return raw


def parse_kml(kml_bytes, label="File"):
    ns = {"kml": KML_NS}
    try:
        root = ET.fromstring(kml_bytes)
    except ET.ParseError as e:
        st.error(f"Gagal parse KML ({label}): {e}")
        return []

    records = []
    for pm in root.findall(".//kml:Placemark", ns):
        name_el = pm.find("kml:name", ns)
        name = name_el.text.strip() if name_el is not None and name_el.text else ""
        ext_data = {}
        schema_data = pm.find(".//kml:SchemaData", ns)
        if schema_data is not None:
            for sd in schema_data.findall("kml:SimpleData", ns):
                k = sd.get("name", "")
                v = sd.text.strip() if sd.text else ""
                ext_data[k] = v
            if not name:
                name = (ext_data.get("NO_SUMUR") or ext_data.get("Name_1")
                        or ext_data.get("name") or "")
        pt = pm.find(".//kml:Point/kml:coordinates", ns)
        if pt is None or not pt.text:
            continue
        parts = pt.text.strip().split(",")
        if len(parts) < 2:
            continue
        try:
            lon, lat = float(parts[0]), float(parts[1])
        except ValueError:
            continue
        records.append({"name": name, "lat": lat, "lon": lon})
    return records


def round_coord(lat, lon, decimal=4):
    """Round coords for grouping"""
    return (round(lat, decimal), round(lon, decimal))


def group_by_coordinate(recs):
    """Group records by coordinate, return dict of coord -> list of records"""
    groups = defaultdict(list)
    for rec in recs:
        key = round_coord(rec["lat"], rec["lon"])
        groups[key].append(rec)
    return groups


def build_grouped_comparison(recs_a, recs_b, threshold_m, label_a, label_b):
    """
    Build comparison with grouped coordinates
    Returns: (df_result, max_dup_a, max_dup_b, stats_a, stats_b)
    """
    
    # Group by coordinate
    groups_a = group_by_coordinate(recs_a)
    groups_b = group_by_coordinate(recs_b)
    
    # Max duplicates
    max_dup_a = max([len(group) for group in groups_a.values()]) if groups_a else 1
    max_dup_b = max([len(group) for group in groups_b.values()]) if groups_b else 1
    
    # Build all coordinate keys
    all_coords = set(list(groups_a.keys()) + list(groups_b.keys()))
    
    # Matching: for each unique coordinate pair, check if within threshold
    rows = []
    overlap_coords = set()
    
    for coord in sorted(all_coords):
        lat, lon = coord
        recs_at_a = groups_a.get(coord, [])
        recs_at_b = groups_b.get(coord, [])
        
        # Check if this coordinate location overlaps
        # If coord exists in both, distance is 0 -> overlap
        is_overlap = len(recs_at_a) > 0 and len(recs_at_b) > 0
        
        if is_overlap:
            overlap_coords.add(coord)
        
        # Build row
        row = {
            "Latitude": lat,
            "Longitude": lon,
        }
        
        # Add File A cols (names + count)
        for i in range(max_dup_a):
            if i < len(recs_at_a):
                row[f"Nama {label_a} - Titik {i+1}"] = recs_at_a[i]["name"]
            else:
                row[f"Nama {label_a} - Titik {i+1}"] = ""
        row[f"Jumlah Sumur {label_a}"] = len(recs_at_a)
        
        # Add File B cols (names + count)
        for i in range(max_dup_b):
            if i < len(recs_at_b):
                row[f"Nama {label_b} - Titik {i+1}"] = recs_at_b[i]["name"]
            else:
                row[f"Nama {label_b} - Titik {i+1}"] = ""
        row[f"Jumlah Sumur {label_b}"] = len(recs_at_b)
        
        # Keterangan (no emoji)
        row["Keterangan"] = "Overlap" if is_overlap else "Tidak Overlap"
        
        rows.append(row)
    
    df = pd.DataFrame(rows)
    
    # Calculate statistics
    stats_a = {
        "total_koordinat": len(groups_a),
        "total_nama": len(recs_a),
        "dup_3": sum(1 for g in groups_a.values() if len(g) >= 3),
        "dup_2": sum(1 for g in groups_a.values() if len(g) == 2),
        "dup_1": sum(1 for g in groups_a.values() if len(g) == 1),
    }
    stats_b = {
        "total_koordinat": len(groups_b),
        "total_nama": len(recs_b),
        "dup_3": sum(1 for g in groups_b.values() if len(g) >= 3),
        "dup_2": sum(1 for g in groups_b.values() if len(g) == 2),
        "dup_1": sum(1 for g in groups_b.values() if len(g) == 1),
    }
    
    return df, max_dup_a, max_dup_b, stats_a, stats_b, overlap_coords


def build_excel_grouped(df_all, df_overlap, df_file_a_only, df_file_b_only, 
                        label_a, label_b, stats_a, stats_b, threshold_m):
    """
    Build Excel dengan format grouped coordinates
    """
    wb = Workbook()
    
    GREEN = "C6EFCE"
    RED = "FFC7CE"
    YELLOW = "FFEB9C"
    BLUE_HDR = "1F4E79"
    GRAY_HDR = "595959"
    WHITE = "FFFFFF"
    LIGHT_GRAY = "F2F2F2"
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    def write_grouped_sheet(ws, df, title, stats_x, stats_y, label_x, label_y, show_stats=True):
        """Write grouped coordinates sheet with statistics (2-column layout)"""
        
        # Title
        ws.merge_cells("A1:M1")
        tc = ws["A1"]
        tc.value = title
        tc.font = Font(bold=True, size=13, color=WHITE)
        tc.fill = PatternFill("solid", fgColor=BLUE_HDR)
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22
        
        # Header
        for ci, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=2, column=ci, value=col_name)
            cell.font = Font(bold=True, size=10, color=WHITE)
            cell.fill = PatternFill("solid", fgColor=GRAY_HDR)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[2].height = 32
        
        # Data
        for ri, row in enumerate(df.itertuples(index=False), start=3):
            row_bg = LIGHT_GRAY if ri % 2 == 0 else WHITE
            for ci, value in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(size=9)
                
                # Color by keterangan
                if "Keterangan" in df.columns and ci == len(df.columns):
                    val_str = str(value) if value else ""
                    if "Overlap" in val_str and "Tidak" not in val_str:
                        cell.fill = PatternFill("solid", fgColor=GREEN)
                        cell.font = Font(size=9, bold=True, color="006100")
                    else:
                        cell.fill = PatternFill("solid", fgColor=RED)
                        cell.font = Font(size=9, bold=True, color="9C0006")
                else:
                    cell.fill = PatternFill("solid", fgColor=row_bg)
        
        # Statistics section (only if show_stats is True)
        if show_stats:
            stat_row = len(df) + 4
            
            # Header: STATISTIK File 1 | File 2
            ws.cell(row=stat_row, column=1, value="STATISTIK").font = Font(bold=True, size=11)
            ws.cell(row=stat_row, column=1).fill = PatternFill("solid", fgColor=BLUE_HDR)
            ws.cell(row=stat_row, column=1).font = Font(bold=True, size=11, color=WHITE)
            
            ws.cell(row=stat_row, column=2, value=label_x).font = Font(bold=True, size=10)
            ws.cell(row=stat_row, column=2).fill = PatternFill("solid", fgColor=LIGHT_GRAY)
            ws.cell(row=stat_row, column=2).alignment = Alignment(horizontal="center")
            
            ws.cell(row=stat_row, column=3, value=label_y).font = Font(bold=True, size=10)
            ws.cell(row=stat_row, column=3).fill = PatternFill("solid", fgColor=LIGHT_GRAY)
            ws.cell(row=stat_row, column=3).alignment = Alignment(horizontal="center")
            
            stat_row += 1
            
            stats_rows = [
                "Total Koordinat dg 3 Nama Sumur",
                "Total Koordinat dg 2 Nama Sumur",
                "Total Titik Koordinat Single",
                "Total Nama Sumur",
                "Total Koordinat",
            ]
            
            stat_keys = ["dup_3", "dup_2", "dup_1", "total_nama", "total_koordinat"]
            
            for label, key in zip(stats_rows, stat_keys):
                # Label
                c1 = ws.cell(row=stat_row, column=1, value=label)
                c1.font = Font(bold=True, size=10)
                c1.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
                c1.border = border
                
                # File 1 value
                val_x = stats_x.get(key, 0)
                c2 = ws.cell(row=stat_row, column=2, value=val_x)
                c2.alignment = Alignment(horizontal="center")
                c2.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
                c2.border = border
                c2.font = Font(size=10)
                
                # File 2 value
                val_y = stats_y.get(key, 0)
                c3 = ws.cell(row=stat_row, column=3, value=val_y)
                c3.alignment = Alignment(horizontal="center")
                c3.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
                c3.border = border
                c3.font = Font(size=10)
                
                stat_row += 1
        
        # Column widths
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        for i in range(4, 20):
            ws.column_dimensions[get_column_letter(i)].width = 16
    
    # Sheet 1: All
    ws_all = wb.active
    ws_all.title = "Semua Data"
    write_grouped_sheet(ws_all, df_all, f"PERBANDINGAN: {label_a} × {label_b}", 
                       stats_a, stats_b, label_a, label_b, show_stats=True)
    
    # Sheet 2: Overlap
    if len(df_overlap) > 0:
        ws_ov = wb.create_sheet("Overlap")
        write_grouped_sheet(ws_ov, df_overlap, f"OVERLAP: {label_a} × {label_b}",
                           stats_a, stats_b, label_a, label_b, show_stats=True)
    
    # Sheet 3: File A only (no stats)
    if len(df_file_a_only) > 0:
        ws_a = wb.create_sheet(f"Hanya {label_a}")
        write_grouped_sheet(ws_a, df_file_a_only, f"HANYA {label_a.upper()}",
                           stats_a, {}, label_a, label_b, show_stats=False)
    
    # Sheet 4: File B only (no stats)
    if len(df_file_b_only) > 0:
        ws_b = wb.create_sheet(f"Hanya {label_b}")
        write_grouped_sheet(ws_b, df_file_b_only, f"HANYA {label_b.upper()}",
                           {}, stats_b, label_a, label_b, show_stats=False)
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── UI ───────────────────────────────────────────────────────────────────────

st.title("🗺️ KML / KMZ Overlap Checker [GROUPED]")
st.caption("Grouped by coordinate dengan adaptive duplicate columns & statistics")

with st.sidebar:
    st.header("⚙️ Pengaturan")
    threshold_m = st.number_input(
        "Threshold jarak overlap (meter)",
        min_value=1, max_value=10_000, value=5, step=1,
    )

col1, col2 = st.columns(2)
with col1:
    st.subheader("📂 File 1")
    file_a = st.file_uploader("Upload KML / KMZ", type=["kml", "kmz"], key="file_a")
    label_a = filename_to_label(file_a) if file_a else "File 1"
    st.text_input("Label", value=label_a, key="label_a_display", disabled=True)

with col2:
    st.subheader("📂 File 2")
    file_b = st.file_uploader("Upload KML / KMZ", type=["kml", "kmz"], key="file_b")
    label_b = filename_to_label(file_b) if file_b else "File 2"
    st.text_input("Label", value=label_b, key="label_b_display", disabled=True)

if st.button("🔍 Proses Perbandingan", type="primary", use_container_width=True):

    if not file_a or not file_b:
        st.warning("Upload dua file dulu.")
        st.stop()

    label_a = filename_to_label(file_a) or "File 1"
    label_b = filename_to_label(file_b) or "File 2"

    with st.spinner("Parsing KML..."):
        kml_bytes_a = extract_kml_bytes(file_a)
        kml_bytes_b = extract_kml_bytes(file_b)
        if kml_bytes_a is None or kml_bytes_b is None:
            st.stop()
        recs_a = parse_kml(kml_bytes_a, label_a)
        recs_b = parse_kml(kml_bytes_b, label_b)

    if not recs_a:
        st.error(f"Tidak ada titik terbaca dari {label_a}.")
        st.stop()
    if not recs_b:
        st.error(f"Tidak ada titik terbaca dari {label_b}.")
        st.stop()

    st.success(f"**{label_a}**: {len(recs_a)} titik  |  **{label_b}**: {len(recs_b)} titik")

    with st.spinner("Building grouped comparison..."):
        df_all, max_dup_a, max_dup_b, stats_a, stats_b, overlap_coords = build_grouped_comparison(
            recs_a, recs_b, threshold_m, label_a, label_b
        )
        
        df_overlap = df_all[df_all["Keterangan"] == "Overlap"].reset_index(drop=True)
        df_file_a_only = df_all[(df_all["Keterangan"] == "Tidak Overlap") & 
                                (df_all[f"Jumlah Sumur {label_a}"] > 0)].reset_index(drop=True)
        df_file_b_only = df_all[(df_all["Keterangan"] == "Tidak Overlap") & 
                                (df_all[f"Jumlah Sumur {label_b}"] > 0)].reset_index(drop=True)

    m1, m2, m3, m4 = st.columns(4)
    m1.metric(f"Koordinat {label_a}", stats_a["total_koordinat"])
    m2.metric(f"Koordinat {label_b}", stats_b["total_koordinat"])
    m3.metric("✅ Overlap", len(df_overlap))
    m4.metric(f"Max Duplikat", f"{max_dup_a}/{max_dup_b}")

    st.markdown("---")

    with st.spinner("Generate Excel..."):
        excel_bytes = build_excel_grouped(
            df_all, df_overlap, df_file_a_only, df_file_b_only,
            label_a, label_b, stats_a, stats_b, threshold_m
        )

    fn = f"Compare_{label_a}_vs_{label_b}.xlsx".replace(" ", "_")
    st.download_button(
        label="📥 Download Excel",
        data=excel_bytes,
        file_name=fn,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary",
    )
