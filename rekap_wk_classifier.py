import streamlit as st
import pandas as pd
import xml.etree.ElementTree as ET
from xml.dom import minidom
from shapely.geometry import Point, Polygon, MultiPolygon, shape
from shapely.ops import unary_union
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import geopandas as gpd
import zipfile
import tempfile
import os
import io
import re

# ─────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="WK Classifier",
    page_icon="🗺️",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ─────────────────────────────────────────────────────────────────────────────
# CSS
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Mono:wght@400;700&family=DM+Sans:wght@300;400;500;600;700&display=swap');

:root {
    --bg: #0d1117; --surface: #161b22; --surface2: #21262d;
    --border: #30363d; --accent: #00d084; --accent2: #ff6b35;
    --text: #e6edf3; --muted: #8b949e; --in-wk: #00d084; --out-wk: #ff6b35;
}
html, body, .stApp { background-color: var(--bg) !important; color: var(--text) !important; font-family: 'DM Sans', sans-serif; }
.main-header { font-family: 'Space Mono', monospace; font-size: 2rem; font-weight: 700; color: var(--text); letter-spacing: -0.02em; margin-bottom: 0.25rem; }
.main-sub { font-size: 0.95rem; color: var(--muted); margin-bottom: 2rem; }
.accent-bar { width: 60px; height: 3px; background: linear-gradient(90deg, var(--accent), var(--accent2)); border-radius: 2px; margin-bottom: 1.5rem; }
.stat-grid { display: grid; grid-template-columns: repeat(3, 1fr); gap: 1rem; margin: 1.5rem 0; }
.stat-box { background: var(--surface); border: 1px solid var(--border); border-radius: 10px; padding: 1.25rem 1rem; text-align: center; }
.stat-box.total { border-top: 3px solid var(--accent2); }
.stat-box.dalam { border-top: 3px solid var(--in-wk); }
.stat-box.luar  { border-top: 3px solid var(--out-wk); }
.stat-num { font-family: 'Space Mono', monospace; font-size: 2.2rem; font-weight: 700; line-height: 1; }
.stat-num.total { color: var(--accent2); } .stat-num.dalam { color: var(--in-wk); } .stat-num.luar { color: var(--out-wk); }
.stat-label { font-size: 0.78rem; color: var(--muted); margin-top: 0.4rem; text-transform: uppercase; letter-spacing: 0.06em; }
.step-row { display: flex; align-items: center; gap: 0.75rem; margin-bottom: 1rem; }
.step-num { width: 28px; height: 28px; border-radius: 50%; background: linear-gradient(135deg, var(--accent), #00a86b); color: #000; font-family: 'Space Mono', monospace; font-size: 0.75rem; font-weight: 700; display: flex; align-items: center; justify-content: center; flex-shrink: 0; }
.step-label { font-size: 0.9rem; font-weight: 500; color: var(--text); }
.file-tag { background: var(--surface2); border: 1px solid var(--border); border-radius: 6px; padding: 0.4rem 0.75rem; font-size: 0.8rem; color: var(--accent); font-family: 'Space Mono', monospace; display: inline-block; margin: 0.2rem; }
.fmt-badge { display: inline-block; padding: 0.15rem 0.5rem; border-radius: 4px; font-size: 0.7rem; font-family: 'Space Mono', monospace; font-weight: 700; margin-left: 6px; }
.fmt-shp  { background: rgba(255,107,53,0.15); color: #ff6b35; border: 1px solid rgba(255,107,53,0.3); }
.fmt-kml  { background: rgba(0,208,132,0.15); color: #00d084; border: 1px solid rgba(0,208,132,0.3); }
.fmt-kmz  { background: rgba(0,150,200,0.15); color: #0096c8; border: 1px solid rgba(0,150,200,0.3); }
.stButton > button { background: linear-gradient(135deg, var(--accent), #00a86b) !important; color: #000 !important; font-family: 'Space Mono', monospace !important; font-weight: 700 !important; font-size: 0.85rem !important; letter-spacing: 0.05em !important; border: none !important; border-radius: 8px !important; width: 100%; }
.stButton > button:hover { opacity: 0.9 !important; transform: translateY(-1px) !important; }
.stDownloadButton > button { background: var(--surface2) !important; color: var(--text) !important; font-family: 'DM Sans', sans-serif !important; border: 1px solid var(--border) !important; border-radius: 8px !important; width: 100%; }
.stDownloadButton > button:hover { border-color: var(--accent) !important; color: var(--accent) !important; }
.stFileUploader > div { background: var(--surface) !important; border: 2px dashed var(--border) !important; border-radius: 10px !important; }
.stFileUploader > div:hover { border-color: var(--accent) !important; }
hr { border-color: var(--border) !important; }
.stTabs [data-baseweb="tab-list"] { background: var(--surface) !important; border-radius: 8px !important; gap: 4px !important; padding: 4px !important; }
.stTabs [data-baseweb="tab"] { background: transparent !important; color: var(--muted) !important; font-family: 'DM Sans', sans-serif !important; border-radius: 6px !important; }
.stTabs [aria-selected="true"] { background: var(--surface2) !important; color: var(--text) !important; }
.section-title { font-family: 'Space Mono', monospace; font-size: 0.72rem; color: var(--muted); text-transform: uppercase; letter-spacing: 0.1em; margin-bottom: 0.5rem; }
#MainMenu, footer, header { visibility: hidden; }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS: PARSING
# ─────────────────────────────────────────────────────────────────────────────

def get_ext(filename):
    return filename.lower().rsplit('.', 1)[-1]


def parse_kml_root(file_bytes, filename):
    """Parse KML/KMZ bytes → XML root."""
    if get_ext(filename) == 'kmz':
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
            kml_name = next((n for n in z.namelist() if n.endswith('.kml')), None)
            if not kml_name:
                raise ValueError("Tidak ada .kml di dalam KMZ")
            kml_bytes = z.read(kml_name)
    else:
        kml_bytes = file_bytes
    return ET.fromstring(kml_bytes)


def shp_bytes_to_geodataframe(uploaded_files):
    """
    Terima list file uploads yg merupakan komponen .shp set
    (.shp, .dbf, .shx, .prj, dll) → GeoDataFrame.
    Bisa juga zip yg berisi shapefile.
    """
    # Cek apakah ada .zip
    zip_files = [f for f in uploaded_files if get_ext(f.name) == 'zip']
    shp_files = [f for f in uploaded_files if get_ext(f.name) == 'shp']

    with tempfile.TemporaryDirectory() as tmpdir:
        if zip_files:
            for zf in zip_files:
                with zipfile.ZipFile(io.BytesIO(zf.read())) as z:
                    z.extractall(tmpdir)
        elif shp_files:
            for f in uploaded_files:
                fpath = os.path.join(tmpdir, f.name)
                with open(fpath, 'wb') as out:
                    out.write(f.read())
        else:
            raise ValueError("Upload file .shp + .dbf + .shx, atau .zip yang berisi shapefile")

        # Cari file .shp di tmpdir
        shp_path = None
        for root, dirs, files in os.walk(tmpdir):
            for fname in files:
                if fname.endswith('.shp'):
                    shp_path = os.path.join(root, fname)
                    break
        if not shp_path:
            raise ValueError("Tidak ditemukan file .shp")

        gdf = gpd.read_file(shp_path)
        # Pastikan CRS WGS84
        if gdf.crs is None:
            gdf = gdf.set_crs('EPSG:4326')
        elif gdf.crs.to_epsg() != 4326:
            gdf = gdf.to_crs('EPSG:4326')
        return gdf


def extract_polygons_from_kml(root):
    """KML root → list of Shapely Polygon."""
    ns = {'kml': 'http://www.opengis.net/kml/2.2'}

    def parse_coords(text):
        pts = []
        for pt in text.strip().split():
            p = pt.split(',')
            if len(p) >= 2:
                try: pts.append((float(p[0]), float(p[1])))
                except: pass
        return pts

    polygons = []
    for coords_el in root.findall('.//kml:Polygon//kml:outerBoundaryIs//kml:coordinates', ns):
        pts = parse_coords(coords_el.text)
        if len(pts) >= 3:
            polygons.append(Polygon(pts))
    # Fallback
    if not polygons:
        for pm in root.findall('.//kml:Placemark', ns):
            if pm.find('.//kml:Point', ns) is None:
                for cel in pm.findall('.//kml:coordinates', ns):
                    pts = parse_coords(cel.text)
                    if len(pts) >= 3:
                        polygons.append(Polygon(pts))
    return polygons


def extract_polygons_from_shp(gdf):
    """GeoDataFrame → list of Shapely Polygon."""
    polys = []
    for geom in gdf.geometry:
        if geom is None: continue
        if geom.geom_type == 'Polygon':
            polys.append(geom)
        elif geom.geom_type == 'MultiPolygon':
            polys.extend(list(geom.geoms))
    return polys


def extract_points_from_kml(root):
    """KML root → DataFrame titik."""
    ns = {'kml': 'http://www.opengis.net/kml/2.2'}
    records = []
    for pm in root.findall('.//kml:Placemark', ns):
        cel = pm.find('.//kml:Point//kml:coordinates', ns)
        if cel is None: continue
        parts = cel.text.strip().split(',')
        if len(parts) < 2: continue
        try:
            lon, lat = float(parts[0]), float(parts[1])
        except: continue
        name_el = pm.find('kml:name', ns)
        desc_el = pm.find('kml:description', ns)
        name = name_el.text.strip() if name_el is not None and name_el.text else ''
        desc = desc_el.text.strip() if desc_el is not None and desc_el.text else ''
        nama_sumur, sumber_info = name, ''
        if ' & ' in name:
            sp = name.split(' & ', 1)
            nama_sumur, sumber_info = sp[0].strip(), sp[1].strip()
        records.append({
            'nama_sumur': nama_sumur, 'sumber_info': sumber_info,
            'nama_lengkap': name, 'deskripsi': desc, 'lon': lon, 'lat': lat,
        })
    return pd.DataFrame(records)


def extract_points_from_shp(gdf):
    """GeoDataFrame Point → DataFrame titik."""
    records = []
    # Kolom kandidat nama
    name_candidates = [c for c in gdf.columns if c.lower() in ['name','nama','label','id','no','kode','sumur']]
    src_candidates  = [c for c in gdf.columns if 'sumber' in c.lower() or 'source' in c.lower() or 'info' in c.lower()]
    name_col = name_candidates[0] if name_candidates else (gdf.columns[0] if len(gdf.columns) > 1 else None)
    src_col  = src_candidates[0]  if src_candidates  else None

    for _, row in gdf.iterrows():
        geom = row.geometry
        if geom is None: continue
        if geom.geom_type == 'Point':
            lon, lat = geom.x, geom.y
        elif geom.geom_type == 'MultiPoint':
            lon, lat = geom.geoms[0].x, geom.geoms[0].y
        else:
            continue
        nama    = str(row[name_col]).strip() if name_col and pd.notna(row[name_col]) else ''
        sumber  = str(row[src_col]).strip()  if src_col  and pd.notna(row[src_col])  else ''
        label   = f"{nama} & {sumber}" if sumber and sumber.lower() != 'nan' else nama
        records.append({
            'nama_sumur': nama, 'sumber_info': sumber,
            'nama_lengkap': label, 'deskripsi': '', 'lon': lon, 'lat': lat,
        })
    return pd.DataFrame(records)


def classify_points(df_points, polygons):
    union_poly = unary_union(polygons)
    df_points = df_points.copy()
    df_points['dalam_wk'] = df_points.apply(
        lambda r: union_poly.contains(Point(r['lon'], r['lat'])), axis=1
    )
    return df_points


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS: OUTPUT BUILDERS
# ─────────────────────────────────────────────────────────────────────────────

def build_kmz_bytes(df, doc_name, icon_color='ff00d084'):
    kml = ET.Element('kml', xmlns='http://www.opengis.net/kml/2.2')
    doc = ET.SubElement(kml, 'Document')
    ET.SubElement(doc, 'name').text = doc_name
    style = ET.SubElement(doc, 'Style', id='s')
    ist = ET.SubElement(style, 'IconStyle')
    ET.SubElement(ist, 'color').text = icon_color
    ET.SubElement(ist, 'scale').text = '0.85'
    icon = ET.SubElement(ist, 'Icon')
    ET.SubElement(icon, 'href').text = 'http://maps.google.com/mapfiles/kml/shapes/donut.png'
    ET.SubElement(ET.SubElement(style, 'LabelStyle'), 'scale').text = '0'

    for _, row in df.iterrows():
        pm = ET.SubElement(doc, 'Placemark')
        ET.SubElement(pm, 'name').text = row['nama_lengkap']
        ET.SubElement(pm, 'styleUrl').text = '#s'
        if row.get('deskripsi', ''):
            ET.SubElement(pm, 'description').text = row['deskripsi']
        pt = ET.SubElement(pm, 'Point')
        ET.SubElement(pt, 'coordinates').text = f"{row['lon']},{row['lat']},0"

    xml_str = minidom.parseString(ET.tostring(kml, encoding='unicode')).toprettyxml(indent='  ')
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr('doc.kml', xml_str.encode('utf-8'))
    return buf.getvalue()


def build_shp_zip_bytes(df, layer_name):
    """DataFrame → ZIP bytes berisi shapefile (.shp/.dbf/.shx/.prj)."""
    gdf = gpd.GeoDataFrame(
        df[['nama_sumur', 'sumber_info', 'lon', 'lat']].copy(),
        geometry=[Point(r['lon'], r['lat']) for _, r in df.iterrows()],
        crs='EPSG:4326'
    )
    gdf = gdf.rename(columns={'nama_sumur': 'nama', 'sumber_info': 'sumber'})

    with tempfile.TemporaryDirectory() as tmpdir:
        shp_path = os.path.join(tmpdir, f"{layer_name}.shp")
        gdf.to_file(shp_path, driver='ESRI Shapefile', encoding='utf-8')

        buf = io.BytesIO()
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            for fname in os.listdir(tmpdir):
                zf.write(os.path.join(tmpdir, fname), fname)
        return buf.getvalue()


def build_excel_bytes(df_all, df_dalam, df_luar, source_names):
    wb = Workbook()

    HDR  = PatternFill('solid', start_color='0D1117')
    HDR2 = PatternFill('solid', start_color='161B22')
    GRN  = PatternFill('solid', start_color='1E8449')
    GRN2 = PatternFill('solid', start_color='27AE60')
    GRNA = PatternFill('solid', start_color='D5F5E3')
    RED  = PatternFill('solid', start_color='922B21')
    RED2 = PatternFill('solid', start_color='C0392B')
    REDA = PatternFill('solid', start_color='FADBD8')
    WHT  = PatternFill('solid', start_color='FFFFFF')

    def thin(c='30363D'):
        return Border(
            left=Side(style='thin', color=c), right=Side(style='thin', color=c),
            top=Side(style='thin', color=c),  bottom=Side(style='thin', color=c))

    HEADERS = ['No.', 'Nama Sumur', 'Sumber Info', 'Longitude', 'Latitude', 'Sumber File']
    WIDTHS  = [6, 20, 24, 18, 18, 20]
    NCOLS   = len(HEADERS)

    def set_cols(ws):
        for col, w in enumerate(WIDTHS, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

    def write_title(ws, row, text, fill, size=12, height=28):
        ws.merge_cells(f'A{row}:{get_column_letter(NCOLS)}{row}')
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(name='Courier New', bold=True, color='FFFFFF', size=size)
        c.fill = fill; c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = height

    def write_headers(ws, row, fill, bc='30363D'):
        for col, h in enumerate(HEADERS, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
            c.fill = fill; c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = thin(bc)
        ws.row_dimensions[row].height = 20

    def write_rows(ws, df, start, alt, bc='30363D'):
        for i, row in df.reset_index(drop=True).iterrows():
            r = start + i; fill = alt if i % 2 == 0 else WHT
            vals = [i+1, row.get('nama_sumur',''), row.get('sumber_info',''),
                    round(float(row['lon']),6), round(float(row['lat']),6),
                    ', '.join(source_names)]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=r, column=col, value=val)
                c.font = Font(name='Arial', size=9); c.fill = fill; c.border = thin(bc)
                c.alignment = Alignment(horizontal='center' if col in [1,4,5] else 'left', vertical='center')
                if col in [4,5]: c.number_format = '0.000000'

    def write_total(ws, row, text, fill, bc='30363D'):
        ws.merge_cells(f'A{row}:{get_column_letter(NCOLS)}{row}')
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(name='Courier New', bold=True, color='FFFFFF', size=10)
        c.fill = fill; c.alignment = Alignment(horizontal='center', vertical='center')
        for col in range(1, NCOLS+1):
            ws.cell(row=row, column=col).fill = fill
            ws.cell(row=row, column=col).border = thin(bc)
        ws.row_dimensions[row].height = 18

    n_dalam = len(df_dalam); n_luar = len(df_luar); total = len(df_all)

    # Sheet 1: Rekap Lengkap
    ws1 = wb.active; ws1.title = 'Rekap Klasifikasi'
    write_title(ws1, 1, 'REKAP KLASIFIKASI TITIK SUMUR – DALAM & LUAR WILAYAH KERJA', HDR, size=12)
    write_title(ws1, 2, f'Total: {total:,}  |  Dalam WK: {n_dalam:,}  |  Luar WK: {n_luar:,}', HDR2, size=10, height=18)
    r = 3
    write_title(ws1, r, f'✅  DALAM WILAYAH KERJA  –  {n_dalam:,} TITIK', GRN, size=11, height=24); r+=1
    write_headers(ws1, r, GRN2, '27AE60'); r+=1
    write_rows(ws1, df_dalam, r, GRNA, 'A9DFBF'); r+=n_dalam
    write_total(ws1, r, f'SUB TOTAL DALAM WK: {n_dalam:,} TITIK', GRN, '27AE60'); r+=2
    write_title(ws1, r, f'❌  LUAR WILAYAH KERJA  –  {n_luar:,} TITIK', RED, size=11, height=24); r+=1
    write_headers(ws1, r, RED2, 'E74C3C'); r+=1
    write_rows(ws1, df_luar, r, REDA, 'F1948A'); r+=n_luar
    write_total(ws1, r, f'SUB TOTAL LUAR WK: {n_luar:,} TITIK', RED, 'E74C3C'); r+=1
    write_total(ws1, r, f'GRAND TOTAL: {total:,} TITIK', HDR); r+=1
    ws1.freeze_panes = 'A3'; set_cols(ws1)

    # Sheet 2: Dalam WK
    ws2 = wb.create_sheet('Dalam WK')
    write_title(ws2, 1, f'TITIK DALAM WILAYAH KERJA  –  {n_dalam:,} TITIK', GRN)
    write_headers(ws2, 2, GRN2, '27AE60')
    write_rows(ws2, df_dalam, 3, GRNA, 'A9DFBF')
    write_total(ws2, n_dalam+3, f'TOTAL: {n_dalam:,} TITIK', GRN, '27AE60')
    ws2.freeze_panes = 'A3'; set_cols(ws2)
    ws2.auto_filter.ref = f'A2:{get_column_letter(NCOLS)}{n_dalam+2}'

    # Sheet 3: Luar WK
    ws3 = wb.create_sheet('Luar WK')
    write_title(ws3, 1, f'TITIK LUAR WILAYAH KERJA  –  {n_luar:,} TITIK', RED)
    write_headers(ws3, 2, RED2, 'E74C3C')
    write_rows(ws3, df_luar, 3, REDA, 'F1948A')
    write_total(ws3, n_luar+3, f'TOTAL: {n_luar:,} TITIK', RED, 'E74C3C')
    ws3.freeze_panes = 'A3'; set_cols(ws3)
    ws3.auto_filter.ref = f'A2:{get_column_letter(NCOLS)}{n_luar+2}'

    # Sheet 4: Summary
    ws4 = wb.create_sheet('Summary')
    ws4.column_dimensions['A'].width = 38; ws4.column_dimensions['B'].width = 20
    write_title(ws4, 1, 'SUMMARY KLASIFIKASI', HDR, size=12)
    sum_rows = [
        ('TOTAL KESELURUHAN', total, HDR),
        ('✅ Dalam Wilayah Kerja', n_dalam, GRN),
        ('   Persentase', f'{n_dalam/total*100:.1f}%', None),
        ('❌ Luar Wilayah Kerja', n_luar, RED),
        ('   Persentase', f'{n_luar/total*100:.1f}%', None),
        ('', '', None),
        ('Sumber File', ', '.join(source_names), None),
    ]
    for i, (label, val, force) in enumerate(sum_rows, 2):
        ca = ws4.cell(row=i, column=1, value=label)
        cb = ws4.cell(row=i, column=2, value=val)
        fill = force if force else (GRNA if i%2==0 else WHT)
        if not label: fill = WHT
        bold = bool(label) and not label.startswith('   ')
        for c in [ca, cb]:
            c.fill = fill; c.border = thin()
            c.font = Font(name='Arial', size=10, bold=bold,
                          color='FFFFFF' if force else '0D1117')
        cb.alignment = Alignment(horizontal='center')

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def fmt_badge(filename):
    ext = get_ext(filename)
    cls = {'shp':'fmt-shp','kml':'fmt-kml','kmz':'fmt-kmz','zip':'fmt-shp'}.get(ext,'fmt-kml')
    return f'<span class="fmt-badge {cls}">.{ext.upper()}</span>'


def detect_file_type(files):
    """Detect if uploaded files are SHP-based or KML/KMZ-based."""
    exts = {get_ext(f.name) for f in files}
    if exts & {'shp','zip'}:
        return 'shp'
    return 'kml'


# ─────────────────────────────────────────────────────────────────────────────
# MAIN UI
# ─────────────────────────────────────────────────────────────────────────────

st.markdown('<div class="main-header">🗺️ WK Classifier</div>', unsafe_allow_html=True)
st.markdown('<div class="accent-bar"></div>', unsafe_allow_html=True)
st.markdown('<div class="main-sub">Klasifikasi titik sumur berdasarkan Wilayah Kerja (WK) · Mendukung KML / KMZ / SHP</div>', unsafe_allow_html=True)

col_left, col_right = st.columns([1, 1.6], gap="large")

with col_left:

    # ── Step 1: Titik Sumur ───────────────────────────────────────────────
    st.markdown('<div class="step-row"><div class="step-num">1</div><div class="step-label">Upload titik sumur</div></div>', unsafe_allow_html=True)
    st.markdown('<div class="section-title">Format: KML · KMZ · SHP (upload .shp+.dbf+.shx) · ZIP shapefile</div>', unsafe_allow_html=True)
    uploaded_points = st.file_uploader(
        "Titik sumur",
        type=['kml', 'kmz', 'shp', 'dbf', 'shx', 'prj', 'cpg', 'zip'],
        accept_multiple_files=True,
        key="points_upload",
        label_visibility="collapsed"
    )
    if uploaded_points:
        for f in uploaded_points:
            st.markdown(f'<span class="file-tag">📍 {f.name}{fmt_badge(f.name)}</span>', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Step 2: Polygon WK ────────────────────────────────────────────────
    st.markdown('<div class="step-row"><div class="step-num">2</div><div class="step-label">Upload polygon WK</div></div>', unsafe_allow_html=True)
    st.markdown('<div class="section-title">Format: KML · KMZ · SHP (upload .shp+.dbf+.shx) · ZIP shapefile</div>', unsafe_allow_html=True)
    uploaded_polygon = st.file_uploader(
        "Polygon WK",
        type=['kml', 'kmz', 'shp', 'dbf', 'shx', 'prj', 'cpg', 'zip'],
        accept_multiple_files=True,
        key="polygon_upload",
        label_visibility="collapsed"
    )
    if uploaded_polygon:
        for f in uploaded_polygon:
            st.markdown(f'<span class="file-tag">🔷 {f.name}{fmt_badge(f.name)}</span>', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Step 3: Run ───────────────────────────────────────────────────────
    st.markdown('<div class="step-row"><div class="step-num">3</div><div class="step-label">Jalankan klasifikasi</div></div>', unsafe_allow_html=True)
    run_btn = st.button("🚀 KLASIFIKASI SEKARANG", use_container_width=True)

# ── Right panel ───────────────────────────────────────────────────────────────
with col_right:
    if run_btn:
        if not uploaded_points:
            st.error("⚠️ Upload minimal 1 file titik sumur")
        elif not uploaded_polygon:
            st.error("⚠️ Upload file polygon WK")
        else:
            with st.spinner("Memproses data..."):
                try:
                    # ── Parse polygon ─────────────────────────────────────
                    poly_type = detect_file_type(uploaded_polygon)
                    if poly_type == 'shp':
                        gdf_poly = shp_bytes_to_geodataframe(uploaded_polygon)
                        polygons  = extract_polygons_from_shp(gdf_poly)
                    else:
                        # KML/KMZ — ambil file pertama
                        pf = uploaded_polygon[0]
                        poly_root = parse_kml_root(pf.read(), pf.name)
                        polygons  = extract_polygons_from_kml(poly_root)

                    if not polygons:
                        st.error("❌ Tidak ditemukan polygon di file WK.")
                        st.stop()

                    # ── Parse titik ───────────────────────────────────────
                    pt_type = detect_file_type(uploaded_points)
                    all_dfs = []
                    source_names = []

                    if pt_type == 'shp':
                        gdf_pts = shp_bytes_to_geodataframe(uploaded_points)
                        df_pt   = extract_points_from_shp(gdf_pts)
                        fname   = next((f.name for f in uploaded_points if get_ext(f.name) in ['shp','zip']), 'shapefile')
                        if not df_pt.empty:
                            df_pt['source_file'] = fname
                            all_dfs.append(df_pt)
                            source_names.append(fname)
                    else:
                        for fp in uploaded_points:
                            if get_ext(fp.name) not in ['kml','kmz']:
                                continue
                            pt_root = parse_kml_root(fp.read(), fp.name)
                            df_pt   = extract_points_from_kml(pt_root)
                            if df_pt.empty:
                                st.warning(f"⚠️ Tidak ada titik di {fp.name}")
                                continue
                            df_pt['source_file'] = fp.name
                            all_dfs.append(df_pt)
                            source_names.append(fp.name)

                    if not all_dfs:
                        st.error("❌ Tidak ada titik yang berhasil dibaca.")
                        st.stop()

                    df_all = pd.concat(all_dfs, ignore_index=True)
                    df_all = classify_points(df_all, polygons)
                    df_dalam = df_all[df_all['dalam_wk']].reset_index(drop=True)
                    df_luar  = df_all[~df_all['dalam_wk']].reset_index(drop=True)

                    st.session_state.update({
                        'df_all': df_all, 'df_dalam': df_dalam,
                        'df_luar': df_luar, 'source_names': source_names, 'done': True
                    })

                except Exception as e:
                    st.error(f"❌ Error: {e}")
                    import traceback
                    st.code(traceback.format_exc())
                    st.stop()

    if st.session_state.get('done'):
        df_all       = st.session_state['df_all']
        df_dalam     = st.session_state['df_dalam']
        df_luar      = st.session_state['df_luar']
        source_names = st.session_state['source_names']
        n_total  = len(df_all)
        n_dalam  = len(df_dalam)
        n_luar   = len(df_luar)

        # Stat boxes
        st.markdown(f"""
        <div class="stat-grid">
            <div class="stat-box total"><div class="stat-num total">{n_total:,}</div><div class="stat-label">Total Titik</div></div>
            <div class="stat-box dalam"><div class="stat-num dalam">{n_dalam:,}</div><div class="stat-label">✅ Dalam WK</div></div>
            <div class="stat-box luar"><div class="stat-num luar">{n_luar:,}</div><div class="stat-label">❌ Luar WK</div></div>
        </div>
        """, unsafe_allow_html=True)

        # Preview tabs
        show_cols = ['nama_sumur','sumber_info','lon','lat']
        col_rename = {'nama_sumur':'Nama Sumur','sumber_info':'Sumber Info','lon':'Longitude','lat':'Latitude'}
        tab1, tab2, tab3 = st.tabs([f"✅ Dalam WK ({n_dalam:,})", f"❌ Luar WK ({n_luar:,})", f"📊 Semua ({n_total:,})"])

        with tab1:
            st.dataframe(df_dalam[show_cols].head(200).rename(columns=col_rename), use_container_width=True, height=260)
            if n_dalam > 200: st.caption(f"Menampilkan 200 dari {n_dalam:,} baris")

        with tab2:
            st.dataframe(df_luar[show_cols].head(200).rename(columns=col_rename), use_container_width=True, height=260)
            if n_luar > 200: st.caption(f"Menampilkan 200 dari {n_luar:,} baris")

        with tab3:
            df_show = df_all[show_cols + ['dalam_wk']].copy()
            df_show['Status'] = df_show['dalam_wk'].map({True:'✅ Dalam WK', False:'❌ Luar WK'})
            st.dataframe(df_show.drop('dalam_wk',axis=1).head(200).rename(columns=col_rename), use_container_width=True, height=260)

        st.markdown("---")
        st.markdown("**⬇️ Download Hasil**")

        # Row 1: Excel + KMZ
        dl1, dl2, dl3 = st.columns(3)
        with dl1:
            excel_bytes = build_excel_bytes(df_all, df_dalam, df_luar, source_names)
            st.download_button("📊 Rekap Excel (.xlsx)", data=excel_bytes,
                file_name="Rekap_Klasifikasi_WK.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
        with dl2:
            st.download_button("🟢 KMZ Dalam WK",
                data=build_kmz_bytes(df_dalam, "Titik Dalam WK", 'ff00d084'),
                file_name="Titik_Dalam_WK.kmz", mime="application/vnd.google-earth.kmz",
                use_container_width=True)
        with dl3:
            st.download_button("🔴 KMZ Luar WK",
                data=build_kmz_bytes(df_luar, "Titik Luar WK", 'ff3555ff'),
                file_name="Titik_Luar_WK.kmz", mime="application/vnd.google-earth.kmz",
                use_container_width=True)

        # Row 2: SHP
        dl4, dl5, _ = st.columns(3)
        with dl4:
            st.download_button("🟩 SHP Dalam WK (.zip)",
                data=build_shp_zip_bytes(df_dalam, "Titik_Dalam_WK"),
                file_name="Titik_Dalam_WK_SHP.zip", mime="application/zip",
                use_container_width=True)
        with dl5:
            st.download_button("🟥 SHP Luar WK (.zip)",
                data=build_shp_zip_bytes(df_luar, "Titik_Luar_WK"),
                file_name="Titik_Luar_WK_SHP.zip", mime="application/zip",
                use_container_width=True)

    elif not run_btn:
        st.markdown("""
        <div style="display:flex; align-items:center; justify-content:center; height:300px; flex-direction:column; gap:1rem; opacity:0.35;">
            <div style="font-size:4rem;">🗺️</div>
            <div style="font-family:'Space Mono',monospace; font-size:0.85rem; color:#8b949e; text-align:center;">
                Upload file & klik Klasifikasi<br>untuk melihat hasil
            </div>
        </div>
        """, unsafe_allow_html=True)
