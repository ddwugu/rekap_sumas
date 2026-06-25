import streamlit as st
import pandas as pd
import xml.etree.ElementTree as ET
from xml.dom import minidom
from shapely.geometry import Point, Polygon
from shapely.ops import unary_union
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import zipfile
import io
import os

# ─────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Multi-Polygon Classifier",
    page_icon="🗺️",
    layout="wide",
)

st.markdown("""
<style>
:root {
    --bg: #0d1117; --surface: #161b22; --surface2: #21262d;
    --border: #30363d; --accent: #00d084; --accent2: #ff6b35;
    --text: #e6edf3; --muted: #8b949e;
}
html, body, .stApp { background-color: var(--bg) !important; color: var(--text) !important; }
.main-header { font-size: 1.8rem; font-weight: 700; margin-bottom: 0.2rem; }
.main-sub { font-size: 0.9rem; color: var(--muted); margin-bottom: 1.5rem; }
.stButton > button { background: linear-gradient(135deg, var(--accent), #00a86b) !important; color:#000 !important; font-weight:700 !important; border:none !important; border-radius:8px !important; width:100%; }
.stDownloadButton > button { background: var(--surface2) !important; color: var(--text) !important; border:1px solid var(--border) !important; border-radius:8px !important; width:100%; }
.poly-slot { background: var(--surface); border: 1px solid var(--border); border-radius: 10px; padding: 1rem; margin-bottom: 0.75rem; }
.stat-grid { display:grid; grid-template-columns: repeat(3,1fr); gap:1rem; margin: 1rem 0; }
.stat-box { background: var(--surface); border:1px solid var(--border); border-radius:10px; padding:1rem; text-align:center; }
.stat-num { font-size:2rem; font-weight:700; }
.stat-label { font-size:0.75rem; color:var(--muted); text-transform:uppercase; }
#MainMenu, footer, header {visibility:hidden;}
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS: PARSING KML/KMZ
# ─────────────────────────────────────────────────────────────────────────────

def get_ext(filename):
    return filename.lower().rsplit('.', 1)[-1]


def parse_kml_root(file_bytes, filename):
    """Parse KML/KMZ bytes -> XML root element"""
    if get_ext(filename) == 'kmz':
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
            kml_name = next((n for n in z.namelist() if n.lower().endswith('.kml')), None)
            if not kml_name:
                raise ValueError("Tidak ada file .kml di dalam KMZ")
            kml_bytes = z.read(kml_name)
    else:
        kml_bytes = file_bytes
    return ET.fromstring(kml_bytes)


def extract_polygons_from_kml(root):
    """KML root -> list of Shapely Polygon"""
    ns = {'kml': 'http://www.opengis.net/kml/2.2'}

    def parse_coords(text):
        pts = []
        for pt in text.strip().split():
            p = pt.split(',')
            if len(p) >= 2:
                try:
                    pts.append((float(p[0]), float(p[1])))
                except:
                    pass
        return pts

    polygons = []
    for coords_el in root.findall('.//kml:Polygon//kml:outerBoundaryIs//kml:coordinates', ns):
        if coords_el.text:
            pts = parse_coords(coords_el.text)
            if len(pts) >= 3:
                polygons.append(Polygon(pts))

    if not polygons:
        for pm in root.findall('.//kml:Placemark', ns):
            if pm.find('.//kml:Point', ns) is None:
                for cel in pm.findall('.//kml:coordinates', ns):
                    if cel.text:
                        pts = parse_coords(cel.text)
                        if len(pts) >= 3:
                            polygons.append(Polygon(pts))
    return polygons


def extract_points_from_kml(root):
    """KML root -> DataFrame titik (nama, lon, lat, deskripsi)"""
    ns = {'kml': 'http://www.opengis.net/kml/2.2'}
    records = []
    for pm in root.findall('.//kml:Placemark', ns):
        cel = pm.find('.//kml:Point//kml:coordinates', ns)
        if cel is None or not cel.text:
            continue
        parts = cel.text.strip().split(',')
        if len(parts) < 2:
            continue
        try:
            lon, lat = float(parts[0]), float(parts[1])
        except:
            continue
        name_el = pm.find('kml:name', ns)
        desc_el = pm.find('kml:description', ns)
        name = name_el.text.strip() if name_el is not None and name_el.text else ''
        desc = desc_el.text.strip() if desc_el is not None and desc_el.text else ''
        records.append({'nama': name, 'lon': lon, 'lat': lat, 'deskripsi': desc})
    return pd.DataFrame(records)


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS: OUTPUT BUILDERS
# ─────────────────────────────────────────────────────────────────────────────

def build_kmz_bytes(df, doc_name, icon_color='ff00d084'):
    kml = ET.Element('kml', xmlns='http://www.opengis.net/kml/2.2')
    doc = ET.SubElement(kml, 'Document')
    ET.SubElement(doc, 'name').text = doc_name
    style = ET.SubElement(doc, 'Style', id='s')
    ist = ET.SubElement(style, 'IconStyle')
    ET.SubElement(ist, 'color').text = icon_color
    ET.SubElement(ist, 'scale').text = '0.85'
    icon = ET.SubElement(ist, 'Icon')
    ET.SubElement(icon, 'href').text = 'http://maps.google.com/mapfiles/kml/shapes/donut.png'
    ET.SubElement(ET.SubElement(style, 'LabelStyle'), 'scale').text = '0'

    for _, row in df.iterrows():
        pm = ET.SubElement(doc, 'Placemark')
        ET.SubElement(pm, 'name').text = str(row['nama']) if row.get('nama') else 'Point'
        ET.SubElement(pm, 'styleUrl').text = '#s'
        pt = ET.SubElement(pm, 'Point')
        ET.SubElement(pt, 'coordinates').text = f"{row['lon']},{row['lat']},0"

    xml_str = minidom.parseString(ET.tostring(kml, encoding='unicode')).toprettyxml(indent='  ')
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr('doc.kml', xml_str.encode('utf-8'))
    return buf.getvalue()


def build_excel_bytes(df_full, df_filtered, polygon_cols, filter_summary):
    """
    df_full     : semua titik + kolom status per polygon
    df_filtered : hasil akhir setelah filter rule diterapkan
    polygon_cols: list nama kolom status polygon, misal ['Polygon 1 (WK_A)', 'Polygon 2 (WK_B)']
    filter_summary: list string ringkasan rule yg dipakai
    """
    wb = Workbook()

    HDR  = PatternFill('solid', start_color='0D1117')
    HDR2 = PatternFill('solid', start_color='161B22')
    GRN  = PatternFill('solid', start_color='1E8449')
    GRN2 = PatternFill('solid', start_color='27AE60')
    BLU  = PatternFill('solid', start_color='2471A3')
    BLU2 = PatternFill('solid', start_color='2E86C1')
    ALT  = PatternFill('solid', start_color='EAF2F8')
    WHT  = PatternFill('solid', start_color='FFFFFF')

    def thin(c='30363D'):
        return Border(left=Side(style='thin', color=c), right=Side(style='thin', color=c),
                       top=Side(style='thin', color=c), bottom=Side(style='thin', color=c))

    base_headers = ['No.', 'Nama Titik', 'Longitude', 'Latitude']
    headers = base_headers + polygon_cols
    ncols = len(headers)

    def set_widths(ws):
        widths = [6, 24, 16, 16] + [22] * len(polygon_cols)
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

    def write_title(ws, row, text, fill, size=12, height=26):
        ws.merge_cells(f'A{row}:{get_column_letter(ncols)}{row}')
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(bold=True, color='FFFFFF', size=size)
        c.fill = fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = height

    def write_headers(ws, row, fill, bc='30363D'):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = Font(bold=True, color='FFFFFF', size=10)
            c.fill = fill
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = thin(bc)
        ws.row_dimensions[row].height = 24

    def write_rows(ws, df, start_row, alt_fill, bc='30363D'):
        for i, row in df.reset_index(drop=True).iterrows():
            r = start_row + i
            fill = alt_fill if i % 2 == 0 else WHT
            vals = [i + 1, row.get('nama', ''), round(float(row['lon']), 6), round(float(row['lat']), 6)]
            for pc in polygon_cols:
                vals.append(row.get(pc, ''))
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=r, column=col, value=val)
                c.font = Font(size=9)
                c.fill = fill
                c.border = thin(bc)
                c.alignment = Alignment(horizontal='center' if col in [1, 3, 4] else 'left', vertical='center')
                if col in [3, 4]:
                    c.number_format = '0.000000'
        return start_row + len(df)

    def write_total(ws, row, text, fill, bc='30363D'):
        ws.merge_cells(f'A{row}:{get_column_letter(ncols)}{row}')
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(bold=True, color='FFFFFF', size=10)
        c.fill = fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        for col in range(1, ncols + 1):
            ws.cell(row=row, column=col).fill = fill
            ws.cell(row=row, column=col).border = thin(bc)
        ws.row_dimensions[row].height = 18

    # Sheet 1: Rekap Lengkap (semua titik vs semua polygon)
    ws1 = wb.active
    ws1.title = 'Rekap Klasifikasi'
    write_title(ws1, 1, f'REKAP KLASIFIKASI TITIK vs {len(polygon_cols)} POLYGON', HDR, size=13)
    write_title(ws1, 2, f'Total titik: {len(df_full):,}', HDR2, size=10, height=18)
    write_headers(ws1, 3, GRN2, '27AE60')
    next_row = write_rows(ws1, df_full, 4, ALT, 'AED6F1')
    write_total(ws1, next_row, f'TOTAL: {len(df_full):,} TITIK', HDR)
    ws1.freeze_panes = 'A4'
    set_widths(ws1)
    ws1.auto_filter.ref = f'A3:{get_column_letter(ncols)}{next_row-1}'

    # Sheet 2: Hasil Filter Final
    ws2 = wb.create_sheet('Hasil Filter Final')
    write_title(ws2, 1, 'HASIL AKHIR SETELAH FILTER', BLU, size=13)
    write_title(ws2, 2, f'Total titik lolos filter: {len(df_filtered):,} dari {len(df_full):,}', HDR2, size=10, height=18)
    write_headers(ws2, 3, BLU2, '2E86C1')
    next_row2 = write_rows(ws2, df_filtered, 4, ALT, 'AED6F1')
    write_total(ws2, next_row2, f'TOTAL LOLOS FILTER: {len(df_filtered):,} TITIK', BLU)
    ws2.freeze_panes = 'A4'
    set_widths(ws2)
    if len(df_filtered) > 0:
        ws2.auto_filter.ref = f'A3:{get_column_letter(ncols)}{next_row2-1}'

    # Sheet 3: Summary Rule Filter
    ws3 = wb.create_sheet('Summary Filter')
    ws3.column_dimensions['A'].width = 45
    ws3.column_dimensions['B'].width = 30
    write_title(ws3, 1, 'RINGKASAN RULE FILTER YANG DIPAKAI', HDR, size=12)
    r = 2
    if not filter_summary:
        ws3.cell(row=r, column=1, value='Tidak ada rule filter aktif (semua polygon hanya info).')
        ws3.cell(row=r, column=1).font = Font(italic=True, size=10)
    else:
        for line in filter_summary:
            c = ws3.cell(row=r, column=1, value=line)
            c.font = Font(size=10)
            c.fill = ALT if r % 2 == 0 else WHT
            c.border = thin()
            ws3.merge_cells(f'A{r}:B{r}')
            r += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# MAIN UI
# ─────────────────────────────────────────────────────────────────────────────

st.markdown('<div class="main-header">🗺️ Multi-Polygon Point Classifier</div>', unsafe_allow_html=True)
st.markdown('<div class="main-sub">Cek titik koordinat terhadap 1–5 polygon, plus filter rule (wajib dalam / wajib luar)</div>', unsafe_allow_html=True)

col_left, col_right = st.columns([1, 1.6], gap="large")

with col_left:

    # ── Upload titik koordinat ──────────────────────────────────────────
    st.markdown("**1️⃣ Upload titik koordinat** (KML/KMZ, 1 file)")
    uploaded_points = st.file_uploader(
        "Titik koordinat", type=['kml', 'kmz'],
        accept_multiple_files=False, key="points_upload",
        label_visibility="collapsed"
    )

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("**2️⃣ Upload Polygon (1 wajib, 2–5 opsional)**")

    polygon_slots = []
    rule_options = ["Hanya info (tidak difilter)", "Wajib DALAM polygon ini", "Wajib LUAR polygon ini"]

    for i in range(1, 6):
        required_tag = " *(wajib)*" if i == 1 else " *(opsional)*"
        with st.container():
            st.markdown(f'<div class="poly-slot">', unsafe_allow_html=True)
            st.markdown(f"**Polygon {i}**{required_tag}")
            pfile = st.file_uploader(f"File Polygon {i}", type=['kml', 'kmz'],
                                      accept_multiple_files=False, key=f"poly_{i}",
                                      label_visibility="collapsed")
            c1, c2 = st.columns([1, 1])
            with c1:
                default_name = os.path.splitext(pfile.name)[0] if pfile else f"Polygon{i}"
                pname = st.text_input(f"Nama Polygon {i}", value=default_name, key=f"pname_{i}",
                                       disabled=(pfile is None))
            with c2:
                prule = st.selectbox(f"Rule Polygon {i}", rule_options, key=f"prule_{i}",
                                      disabled=(pfile is None))
            st.markdown('</div>', unsafe_allow_html=True)
            polygon_slots.append({'idx': i, 'file': pfile, 'name': pname, 'rule': prule})

    st.markdown("<br>", unsafe_allow_html=True)
    run_btn = st.button("🚀 JALANKAN KLASIFIKASI & FILTER", use_container_width=True)

with col_right:
    if run_btn:
        if uploaded_points is None:
            st.error("⚠️ Upload file titik koordinat dulu.")
            st.stop()
        if polygon_slots[0]['file'] is None:
            st.error("⚠️ Polygon 1 wajib diupload (minimal 1 polygon).")
            st.stop()

        with st.spinner("Memproses..."):
            try:
                # Parse titik
                pts_root = parse_kml_root(uploaded_points.read(), uploaded_points.name)
                df_pts = extract_points_from_kml(pts_root)

                if df_pts.empty:
                    st.error("❌ Tidak ada titik yang berhasil dibaca dari file KML/KMZ.")
                    st.stop()

                active_slots = [s for s in polygon_slots if s['file'] is not None]
                polygon_cols = []
                rule_map = {}  # col_name -> ('dalam'|'luar')
                filter_summary = []

                for slot in active_slots:
                    poly_root = parse_kml_root(slot['file'].read(), slot['file'].name)
                    polys = extract_polygons_from_kml(poly_root)
                    if not polys:
                        st.warning(f"⚠️ Tidak ditemukan polygon di file Polygon {slot['idx']} ({slot['file'].name}), dilewati.")
                        continue
                    union_poly = unary_union(polys)

                    col_name = f"Polygon {slot['idx']} ({slot['name']})"
                    df_pts[col_name] = df_pts.apply(
                        lambda r: 'Dalam' if union_poly.contains(Point(r['lon'], r['lat'])) else 'Luar',
                        axis=1
                    )
                    polygon_cols.append(col_name)

                    if slot['rule'] == "Wajib DALAM polygon ini":
                        rule_map[col_name] = 'Dalam'
                        filter_summary.append(f"{col_name} → wajib DALAM")
                    elif slot['rule'] == "Wajib LUAR polygon ini":
                        rule_map[col_name] = 'Luar'
                        filter_summary.append(f"{col_name} → wajib LUAR")

                if not polygon_cols:
                    st.error("❌ Tidak ada polygon valid yang berhasil diproses.")
                    st.stop()

                # Terapkan filter rule (AND di semua polygon yang punya rule aktif)
                df_filtered = df_pts.copy()
                for col_name, required_status in rule_map.items():
                    df_filtered = df_filtered[df_filtered[col_name] == required_status]
                df_filtered = df_filtered.reset_index(drop=True)

                st.session_state.update({
                    'df_pts': df_pts,
                    'df_filtered': df_filtered,
                    'polygon_cols': polygon_cols,
                    'filter_summary': filter_summary,
                    'done': True,
                })

            except Exception as e:
                st.error(f"❌ Error: {e}")
                import traceback
                st.code(traceback.format_exc())
                st.stop()

    if st.session_state.get('done'):
        df_pts = st.session_state['df_pts']
        df_filtered = st.session_state['df_filtered']
        polygon_cols = st.session_state['polygon_cols']
        filter_summary = st.session_state['filter_summary']

        n_total = len(df_pts)
        n_filtered = len(df_filtered)

        st.markdown(f"""
        <div class="stat-grid">
            <div class="stat-box"><div class="stat-num">{n_total:,}</div><div class="stat-label">Total Titik</div></div>
            <div class="stat-box"><div class="stat-num">{len(polygon_cols)}</div><div class="stat-label">Polygon Diproses</div></div>
            <div class="stat-box"><div class="stat-num">{n_filtered:,}</div><div class="stat-label">Lolos Filter</div></div>
        </div>
        """, unsafe_allow_html=True)

        if filter_summary:
            st.info("**Rule filter aktif:**\n" + "\n".join(f"- {s}" for s in filter_summary))
        else:
            st.info("Tidak ada rule filter aktif — semua polygon hanya untuk info, hasil filter = semua titik.")

        tab1, tab2 = st.tabs([f"📊 Rekap Semua Titik ({n_total:,})", f"✅ Hasil Filter Final ({n_filtered:,})"])

        show_cols = ['nama', 'lon', 'lat'] + polygon_cols
        rename_map = {'nama': 'Nama Titik', 'lon': 'Longitude', 'lat': 'Latitude'}

        with tab1:
            st.dataframe(df_pts[show_cols].rename(columns=rename_map), use_container_width=True, height=320)

        with tab2:
            if n_filtered > 0:
                st.dataframe(df_filtered[show_cols].rename(columns=rename_map), use_container_width=True, height=320)
            else:
                st.warning("Tidak ada titik yang lolos semua rule filter.")

        st.markdown("---")
        st.markdown("**⬇️ Download Hasil**")

        dl1, dl2, dl3 = st.columns(3)
        with dl1:
            excel_bytes = build_excel_bytes(df_pts, df_filtered, polygon_cols, filter_summary)
            st.download_button("📊 Excel Rekap + Filter", data=excel_bytes,
                                file_name="Rekap_Klasifikasi_Multi_Polygon.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True)
        with dl2:
            st.download_button("🟢 KMZ Hasil Filter Final",
                                data=build_kmz_bytes(df_filtered, "Hasil Filter Final", 'ff00d084'),
                                file_name="Hasil_Filter_Final.kmz",
                                mime="application/vnd.google-earth.kmz",
                                use_container_width=True,
                                disabled=(n_filtered == 0))
        with dl3:
            st.download_button("🗺️ KMZ Semua Titik (rekap)",
                                data=build_kmz_bytes(df_pts, "Semua Titik", 'ff3555ff'),
                                file_name="Semua_Titik_Rekap.kmz",
                                mime="application/vnd.google-earth.kmz",
                                use_container_width=True)

    elif not run_btn:
        st.markdown("""
        <div style="display:flex; align-items:center; justify-content:center; height:280px; flex-direction:column; gap:1rem; opacity:0.35;">
            <div style="font-size:3.5rem;">🗺️</div>
            <div style="text-align:center; color:#8b949e;">Upload titik & minimal 1 polygon,<br>lalu klik Jalankan Klasifikasi</div>
        </div>
        """, unsafe_allow_html=True)
