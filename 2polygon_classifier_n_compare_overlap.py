import streamlit as st
import pandas as pd
import zipfile
import io
import os
from xml.etree import ElementTree as ET
from math import radians, cos, sin, asin, sqrt
from collections import defaultdict
from shapely.geometry import Point, Polygon
from shapely.ops import unary_union
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="KML/KMZ Compare + Polygon Classifier", page_icon="🗺️", layout="wide")

KML_NS = "http://www.opengis.net/kml/2.2"


# ── HELPERS ──────────────────────────────────────────────────────────────────

def filename_to_label(uploaded_file):
    if uploaded_file is None:
        return ""
    base = os.path.splitext(uploaded_file.name)[0]
    return base.replace("_", " ").replace("-", " ")


def haversine_m(lat1, lon1, lat2, lon2):
    R = 6_371_000
    phi1, phi2 = radians(lat1), radians(lat2)
    a = sin((phi2 - phi1) / 2) ** 2 + cos(phi1) * cos(phi2) * sin((radians(lon2 - lon1)) / 2) ** 2
    return 2 * R * asin(sqrt(a))


def extract_digits(name):
    return ''.join(c for c in name if c.isdigit())


def compare_names_by_digits(name_a, name_b):
    digits_a = extract_digits(name_a)
    digits_b = extract_digits(name_b)

    if not digits_a or not digits_b:
        return "Tidak Ada Digit"

    if len(digits_a) >= 4 and len(digits_b) >= 4:
        if digits_a[-4:] == digits_b[-4:]:
            return "Sama"

    if digits_a == digits_b:
        return "Sama"

    for i in range(len(digits_a) - 3):
        if digits_a[i:i+4] in digits_b:
            return "Sama"

    return "Berbeda"


def extract_kml_bytes(uploaded_file):
    name = uploaded_file.name.lower()
    raw = uploaded_file.read()
    if name.endswith(".kmz"):
        with zipfile.ZipFile(io.BytesIO(raw)) as z:
            kml_names = [n for n in z.namelist() if n.lower().endswith(".kml")]
            if not kml_names:
                st.error(f"Tidak ada file .kml di dalam {uploaded_file.name}")
                return None
            return z.read(kml_names[0])
    return raw


def parse_kml(kml_bytes, label="File"):
    """
    Parse titik KML → list dict.
    lat_str/lon_str = STRING ASLI dari file → dipakai di output (full precision, zero rounding).
    """
    ns = {"kml": KML_NS}
    try:
        root = ET.fromstring(kml_bytes)
    except ET.ParseError as e:
        st.error(f"Gagal parse KML ({label}): {e}")
        return []

    records = []
    for pm in root.findall(".//kml:Placemark", ns):
        name_el = pm.find("kml:name", ns)
        name = name_el.text.strip() if name_el is not None and name_el.text else ""
        ext_data = {}
        schema_data = pm.find(".//kml:SchemaData", ns)
        if schema_data is not None:
            for sd in schema_data.findall("kml:SimpleData", ns):
                k = sd.get("name", "")
                v = sd.text.strip() if sd.text else ""
                ext_data[k] = v
            if not name:
                name = (ext_data.get("NO_SUMUR") or ext_data.get("Name_1")
                        or ext_data.get("name") or "")
        pt = pm.find(".//kml:Point/kml:coordinates", ns)
        if pt is None or not pt.text:
            continue
        parts = pt.text.strip().split(",")
        if len(parts) < 2:
            continue
        lon_str = parts[0].strip()
        lat_str = parts[1].strip()
        try:
            lon, lat = float(lon_str), float(lat_str)
        except ValueError:
            continue
        records.append({"name": name, "lat": lat, "lon": lon,
                        "lat_str": lat_str, "lon_str": lon_str,
                        "order": len(records)})
    return records


def extract_polygons_from_kml(kml_bytes):
    """Extract Shapely Polygon dari KML bytes"""
    ns = {'kml': KML_NS}
    root = ET.fromstring(kml_bytes)

    def parse_coords(text):
        pts = []
        for pt in text.strip().split():
            p = pt.split(',')
            if len(p) >= 2:
                try:
                    pts.append((float(p[0]), float(p[1])))
                except:
                    pass
        return pts

    polygons = []
    for coords_el in root.findall('.//kml:Polygon//kml:outerBoundaryIs//kml:coordinates', ns):
        if coords_el.text:
            pts = parse_coords(coords_el.text)
            if len(pts) >= 3:
                polygons.append(Polygon(pts))

    if not polygons:
        for pm in root.findall('.//kml:Placemark', ns):
            if pm.find('.//kml:Point', ns) is None:
                for cel in pm.findall('.//kml:coordinates', ns):
                    if cel.text:
                        pts = parse_coords(cel.text)
                        if len(pts) >= 3:
                            polygons.append(Polygon(pts))
    return polygons


# ── CLUSTERING & MATCHING ────────────────────────────────────────────────────

def _grid_cell(lat, lon, cell_deg):
    return (int(lat // cell_deg), int(lon // cell_deg))


def cluster_points(recs, threshold_m):
    """
    Kelompokkan titik duplikat dalam SATU file jadi satu 'lokasi'.

    threshold_m = 0 (DEFAULT) → HANYA titik dengan koordinat IDENTIK PERSIS
    (string sama) yang digabung. Sumur berbeda yang jaraknya dekat (mis. 3–4 m)
    TETAP jadi lokasi terpisah, masing-masing dengan koordinat aslinya.

    threshold_m > 0 → titik berjarak ≤ threshold digabung (haversine + grid index).
    """
    if threshold_m <= 0:
        groups = defaultdict(list)
        for rec in recs:
            groups[(rec["lat_str"], rec["lon_str"])].append(rec)
        return [{
            "lat": g[0]["lat"], "lon": g[0]["lon"],
            "lat_str": g[0]["lat_str"], "lon_str": g[0]["lon_str"],
            "recs": g,
        } for g in groups.values()]

    cell_deg = max(threshold_m / 111320.0, 1e-9)
    grid = defaultdict(list)
    clusters = []

    for rec in recs:
        ci, cj = _grid_cell(rec["lat"], rec["lon"], cell_deg)
        found = None
        for di in (-1, 0, 1):
            for dj in (-1, 0, 1):
                for k in grid.get((ci + di, cj + dj), []):
                    c = clusters[k]
                    if haversine_m(rec["lat"], rec["lon"], c["lat"], c["lon"]) <= threshold_m:
                        found = k
                        break
                if found is not None:
                    break
            if found is not None:
                break

        if found is not None:
            clusters[found]["recs"].append(rec)
        else:
            clusters.append({
                "lat": rec["lat"], "lon": rec["lon"],
                "lat_str": rec["lat_str"], "lon_str": rec["lon_str"],
                "recs": [rec],
            })
            grid[(ci, cj)].append(len(clusters) - 1)

    return clusters


def match_clusters(clusters_a, clusters_b, threshold_m):
    """
    Greedy 1-to-1 matching antar lokasi A dan B berdasarkan jarak haversine,
    pair TERDEKAT diprioritaskan. Overlap = jarak ≤ threshold_m.
    """
    cell_deg = max(threshold_m / 111320.0, 1e-9)
    grid_b = defaultdict(list)
    for j, c in enumerate(clusters_b):
        grid_b[_grid_cell(c["lat"], c["lon"], cell_deg)].append(j)

    candidates = []
    for i, ca in enumerate(clusters_a):
        ci, cj = _grid_cell(ca["lat"], ca["lon"], cell_deg)
        for di in (-1, 0, 1):
            for dj in (-1, 0, 1):
                for j in grid_b.get((ci + di, cj + dj), []):
                    cb = clusters_b[j]
                    d = haversine_m(ca["lat"], ca["lon"], cb["lat"], cb["lon"])
                    if d <= threshold_m:
                        candidates.append((d, i, j))

    candidates.sort()
    matched_a, matched_b, matches = set(), set(), []
    for d, i, j in candidates:
        if i in matched_a or j in matched_b:
            continue
        matched_a.add(i)
        matched_b.add(j)
        matches.append((i, j, d))

    return matches, matched_a, matched_b


def _cluster_stats(clusters, recs):
    return {
        "total_koordinat": len(clusters),
        "total_nama": len(recs),
        "dup_3": sum(1 for c in clusters if len(c["recs"]) >= 3),
        "dup_2": sum(1 for c in clusters if len(c["recs"]) == 2),
        "dup_1": sum(1 for c in clusters if len(c["recs"]) == 1),
    }


def build_grouped_comparison(recs_a, recs_b, threshold_m, dup_threshold_m, label_a, label_b):
    """
    - Duplikat intra-file: dup_threshold_m (default 0 = hanya koordinat identik persis).
    - Overlap antar file: jarak haversine ≤ threshold_m.
    - Koordinat A dan B DITAMPILKAN TERPISAH, string asli file (full precision).
    """
    clusters_a = cluster_points(recs_a, dup_threshold_m)
    clusters_b = cluster_points(recs_b, dup_threshold_m)

    matches, matched_a, matched_b = match_clusters(clusters_a, clusters_b, threshold_m)

    max_dup_a = max([len(c["recs"]) for c in clusters_a], default=1)
    max_dup_b = max([len(c["recs"]) for c in clusters_b], default=1)

    col_lat_a = f"Latitude {label_a}"
    col_lon_a = f"Longitude {label_a}"
    col_lat_b = f"Latitude {label_b}"
    col_lon_b = f"Longitude {label_b}"

    rows = []

    def make_row(ca, cb, dist_m):
        row = {}
        # Koordinat A & B masing-masing, string asli file → jelas terlihat bedanya
        row[col_lat_a] = ca["lat_str"] if ca is not None else ""
        row[col_lon_a] = ca["lon_str"] if ca is not None else ""
        row[col_lat_b] = cb["lat_str"] if cb is not None else ""
        row[col_lon_b] = cb["lon_str"] if cb is not None else ""

        recs_at_a = ca["recs"] if ca is not None else []
        recs_at_b = cb["recs"] if cb is not None else []

        for i in range(max_dup_a):
            row[f"Nama {label_a} - Titik {i+1}"] = recs_at_a[i]["name"] if i < len(recs_at_a) else ""
        row[f"Jumlah Sumur {label_a}"] = len(recs_at_a)

        for i in range(max_dup_b):
            row[f"Nama {label_b} - Titik {i+1}"] = recs_at_b[i]["name"] if i < len(recs_at_b) else ""
        row[f"Jumlah Sumur {label_b}"] = len(recs_at_b)

        if recs_at_a and recs_at_b:
            row["Kesamaan Nama"] = compare_names_by_digits(recs_at_a[0]["name"], recs_at_b[0]["name"])
        else:
            row["Kesamaan Nama"] = "-"

        row["Jarak (m)"] = round(dist_m, 2) if dist_m is not None else ""
        row["Keterangan"] = "Overlap" if (recs_at_a and recs_at_b) else "Tidak Overlap"

        src = ca if ca is not None else cb
        row["_lat"] = src["lat"]
        row["_lon"] = src["lon"]
        return row

    for i, j, d in matches:
        rows.append(make_row(clusters_a[i], clusters_b[j], d))

    for i, ca in enumerate(clusters_a):
        if i not in matched_a:
            rows.append(make_row(ca, None, None))

    for j, cb in enumerate(clusters_b):
        if j not in matched_b:
            rows.append(make_row(None, cb, None))

    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values(["_lat", "_lon"]).reset_index(drop=True)

    stats_a = _cluster_stats(clusters_a, recs_a)
    stats_b = _cluster_stats(clusters_b, recs_b)

    coord_cols = [col_lat_a, col_lon_a, col_lat_b, col_lon_b]

    return (df, max_dup_a, max_dup_b, stats_a, stats_b, coord_cols,
            clusters_a, clusters_b, matches)


# ── PER SUMUR (1 BARIS = 1 NAMA SUMUR, URUTAN ASLI FILE TERBANYAK) ───────────

def _map_rec_to_cluster(clusters):
    m = {}
    for idx, c in enumerate(clusters):
        for rec in c["recs"]:
            m[id(rec)] = idx
    return m


def build_per_row_comparison(recs_a, recs_b, clusters_a, clusters_b, matches,
                             label_a, label_b, primary="auto"):
    """
    Sheet 'Per Sumur':
    - Baris mengikuti file ACUAN (primary). File acuan ditentukan `primary`:
        "auto" → file dengan JUMLAH NAMA TERBANYAK (perilaku lama, default)
        "a"    → selalu ikut File 1 (label_a)
        "b"    → selalu ikut File 2 (label_b)
    - Urutan baris = urutan ASLI file acuan.
    - Setiap nama sumur = 1 baris, walau koordinatnya duplikat/redundant.
      Contoh: x1 & x2 di koordinat sama → tetap 2 baris terpisah.
    - Tiap baris dipasangkan dengan lokasi match di file satunya (secondary):
      semua nama di lokasi itu (y1, y2, y3) tampil di kolom perbandingan,
      lengkap koordinat, kesamaan nama, jarak, keterangan.
    """
    # Tentukan file acuan (primary) sesuai pilihan user
    if primary == "a":
        use_a = True
    elif primary == "b":
        use_a = False
    else:  # "auto" → file terbanyak
        use_a = len(recs_a) >= len(recs_b)

    if use_a:
        recs_p, clusters_p = recs_a, clusters_a
        clusters_s = clusters_b
        label_p, label_s = label_a, label_b
        match_map = {i: (j, d) for i, j, d in matches}
    else:
        recs_p, clusters_p = recs_b, clusters_b
        clusters_s = clusters_a
        label_p, label_s = label_b, label_a
        match_map = {j: (i, d) for i, j, d in matches}

    col_lat_p = f"Latitude {label_p}"
    col_lon_p = f"Longitude {label_p}"
    col_lat_s = f"Latitude {label_s}"
    col_lon_s = f"Longitude {label_s}"

    # Urutkan cluster berdasarkan kemunculan pertama di file (urutan asli),
    # anggota redundant dalam 1 cluster ditulis BERURUTAN (dikelompokkan).
    ordered_clusters = sorted(
        range(len(clusters_p)),
        key=lambda ci: min(r["order"] for r in clusters_p[ci]["recs"])
    )

    rows = []
    no = 0
    for grp_id, ci in enumerate(ordered_clusters, start=1):
        cluster = clusters_p[ci]
        member_recs = sorted(cluster["recs"], key=lambda r: r["order"])
        is_redundant = len(member_recs) > 1
        pair = match_map.get(ci)

        for rec in member_recs:
            no += 1
            row = {"No": no, f"Nama {label_p}": rec["name"],
                   col_lat_p: rec["lat_str"], col_lon_p: rec["lon_str"],
                   f"Jumlah Redundant {label_p}": len(member_recs)}

            if pair is not None:
                cs = clusters_s[pair[0]]
                names_s = [r["name"] for r in cs["recs"]]
                row[f"Nama {label_s}"] = ", ".join(names_s)
                row[f"Jumlah Sumur {label_s}"] = len(names_s)
                row[col_lat_s] = cs["lat_str"]
                row[col_lon_s] = cs["lon_str"]
                row["Kesamaan Nama"] = compare_names_by_digits(rec["name"], names_s[0])
                row["Jarak (m)"] = round(pair[1], 2)
                row["Keterangan"] = "Overlap"
            else:
                row[f"Nama {label_s}"] = ""
                row[f"Jumlah Sumur {label_s}"] = 0
                row[col_lat_s] = ""
                row[col_lon_s] = ""
                row["Kesamaan Nama"] = "-"
                row["Jarak (m)"] = ""
                row["Keterangan"] = "Tidak Overlap"

            row["_lat"] = rec["lat"]
            row["_lon"] = rec["lon"]
            row["_grup"] = grp_id
            row["_redundant"] = is_redundant
            rows.append(row)

    df = pd.DataFrame(rows)
    return df, label_p, label_s


def per_row_fills(df_per_row):
    """
    Warna baris sheet Per Sumur: kelompok redundant diberi warna biru,
    2 shade biru bergantian antar grup redundant yang bersebelahan
    biar batas antar kelompok keliatan jelas.
    """
    BLUE_1 = "DDEBF7"
    BLUE_2 = "B4C6E7"
    fills = []
    shade_toggle = 0
    prev_grup = None
    for _, r in df_per_row.iterrows():
        if r["_redundant"]:
            if r["_grup"] != prev_grup:
                shade_toggle ^= 1
            fills.append(BLUE_1 if shade_toggle else BLUE_2)
            prev_grup = r["_grup"]
        else:
            fills.append(None)
            prev_grup = None
    return fills


def build_excel_grouped(df_all, df_overlap, df_file_a_only, df_file_b_only, df_lolos_spasial,
                        df_per_row, per_row_label_p,
                        label_a, label_b, stats_a, stats_b, threshold_m, dup_threshold_m,
                        coord_cols, polygon_cols=None):
    wb = Workbook()

    GREEN = "C6EFCE"
    RED = "FFC7CE"
    BLUE_HDR = "1F4E79"
    GRAY_HDR = "595959"
    WHITE = "FFFFFF"
    LIGHT_GRAY = "F2F2F2"
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    if polygon_cols is None:
        polygon_cols = []

    def visible_df(df):
        return df[[c for c in df.columns if not c.startswith("_")]]

    def write_grouped_sheet(ws, df_raw, title, stats_x, stats_y, label_x, label_y,
                            show_stats=True, extra_stats=None, row_fills=None):
        df = visible_df(df_raw)
        ws.merge_cells("A1:M1")
        tc = ws["A1"]
        tc.value = title
        tc.font = Font(bold=True, size=13, color=WHITE)
        tc.fill = PatternFill("solid", fgColor=BLUE_HDR)
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        for ci, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=2, column=ci, value=col_name)
            cell.font = Font(bold=True, size=10, color=WHITE)
            cell.fill = PatternFill("solid", fgColor=GRAY_HDR)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[2].height = 32

        for ri, row in enumerate(df.itertuples(index=False), start=3):
            override = row_fills[ri - 3] if row_fills is not None else None
            row_bg = override if override else (LIGHT_GRAY if ri % 2 == 0 else WHITE)
            for ci, value in enumerate(row, 1):
                col_name = df.columns[ci - 1]

                # Koordinat: ANGKA full precision (tanpa round), tampil s.d. 10 desimal
                if col_name in coord_cols:
                    if value != "" and value is not None:
                        try:
                            cell = ws.cell(row=ri, column=ci, value=float(value))
                            cell.number_format = '0.##########'
                        except (ValueError, TypeError):
                            cell = ws.cell(row=ri, column=ci, value=value)
                    else:
                        cell = ws.cell(row=ri, column=ci, value="")
                else:
                    cell = ws.cell(row=ri, column=ci, value=value)

                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(size=9)

                if col_name == "Keterangan":
                    val_str = str(value) if value else ""
                    if "Overlap" in val_str and "Tidak" not in val_str:
                        cell.fill = PatternFill("solid", fgColor=GREEN)
                        cell.font = Font(size=9, bold=True, color="006100")
                    else:
                        cell.fill = PatternFill("solid", fgColor=RED)
                        cell.font = Font(size=9, bold=True, color="9C0006")
                elif col_name == "Status Analisa Spasial":
                    val_str = str(value) if value else ""
                    if val_str == "Lolos":
                        cell.fill = PatternFill("solid", fgColor=GREEN)
                        cell.font = Font(size=9, bold=True, color="006100")
                    elif val_str == "Tidak Lolos":
                        cell.fill = PatternFill("solid", fgColor=RED)
                        cell.font = Font(size=9, bold=True, color="9C0006")
                    else:
                        cell.fill = PatternFill("solid", fgColor=row_bg)
                else:
                    cell.fill = PatternFill("solid", fgColor=row_bg)

        if show_stats:
            stat_row = len(df) + 4

            ws.cell(row=stat_row, column=1, value="STATISTIK").font = Font(bold=True, size=11, color=WHITE)
            ws.cell(row=stat_row, column=1).fill = PatternFill("solid", fgColor=BLUE_HDR)

            ws.cell(row=stat_row, column=2, value=label_x).font = Font(bold=True, size=10)
            ws.cell(row=stat_row, column=2).fill = PatternFill("solid", fgColor=LIGHT_GRAY)
            ws.cell(row=stat_row, column=2).alignment = Alignment(horizontal="center")

            ws.cell(row=stat_row, column=3, value=label_y).font = Font(bold=True, size=10)
            ws.cell(row=stat_row, column=3).fill = PatternFill("solid", fgColor=LIGHT_GRAY)
            ws.cell(row=stat_row, column=3).alignment = Alignment(horizontal="center")

            stat_row += 1

            stats_rows = [
                "Total Lokasi dg ≥3 Nama Sumur",
                "Total Lokasi dg 2 Nama Sumur",
                "Total Lokasi Single",
                "Total Nama Sumur",
                "Total Lokasi Unik",
            ]
            stat_keys = ["dup_3", "dup_2", "dup_1", "total_nama", "total_koordinat"]

            for label, key in zip(stats_rows, stat_keys):
                c1 = ws.cell(row=stat_row, column=1, value=label)
                c1.font = Font(bold=True, size=10)
                c1.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
                c1.border = border

                c2 = ws.cell(row=stat_row, column=2, value=stats_x.get(key, 0))
                c2.alignment = Alignment(horizontal="center")
                c2.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
                c2.border = border
                c2.font = Font(size=10)

                c3 = ws.cell(row=stat_row, column=3, value=stats_y.get(key, 0))
                c3.alignment = Alignment(horizontal="center")
                c3.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
                c3.border = border
                c3.font = Font(size=10)

                stat_row += 1

            for label, val in [("Threshold Overlap antar file (meter)", threshold_m),
                               ("Threshold Duplikat dalam file (meter)",
                                dup_threshold_m if dup_threshold_m > 0 else "0 (identik persis)")]:
                c1 = ws.cell(row=stat_row, column=1, value=label)
                c1.font = Font(bold=True, size=10)
                c1.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
                c1.border = border
                c3 = ws.cell(row=stat_row, column=3, value=val)
                c3.alignment = Alignment(horizontal="center")
                c3.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
                c3.border = border
                c3.font = Font(size=10)
                stat_row += 1

            if extra_stats:
                for label, val in extra_stats:
                    c1 = ws.cell(row=stat_row, column=1, value=label)
                    c1.font = Font(bold=True, size=10)
                    c1.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
                    c1.border = border

                    c2 = ws.cell(row=stat_row, column=2, value="")
                    c2.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
                    c2.border = border

                    c3 = ws.cell(row=stat_row, column=3, value=val)
                    c3.alignment = Alignment(horizontal="center")
                    c3.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
                    c3.border = border
                    c3.font = Font(size=10)

                    stat_row += 1

        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 16
        for i in range(5, 32):
            ws.column_dimensions[get_column_letter(i)].width = 16

    # Sheet 1: Semua Data
    ws_all = wb.active
    ws_all.title = "Semua Data"
    write_grouped_sheet(ws_all, df_all,
                        f"PERBANDINGAN: {label_a} × {label_b} (overlap ≤ {threshold_m} m)",
                        stats_a, stats_b, label_a, label_b, show_stats=True)

    # Sheet 2: Per Sumur (baris = tiap nama sumur file acuan, urutan asli)
    if df_per_row is not None and len(df_per_row) > 0:
        ws_pr = wb.create_sheet("Per Sumur")
        n_overlap_pr = len(df_per_row[df_per_row["Keterangan"] == "Overlap"])
        n_redundant_pr = int(df_per_row["_redundant"].sum()) if "_redundant" in df_per_row.columns else 0
        extra_stats_pr = [
            ("Total Baris (nama sumur " + per_row_label_p + ")", len(df_per_row)),
            ("Baris Overlap", n_overlap_pr),
            ("Baris Tidak Overlap", len(df_per_row) - n_overlap_pr),
            ("Baris Redundant (biru)", n_redundant_pr),
        ]
        if "Status Analisa Spasial" in df_per_row.columns:
            extra_stats_pr.append(("Lolos Analisa Spasial",
                                   len(df_per_row[df_per_row["Status Analisa Spasial"] == "Lolos"])))
            extra_stats_pr.append(("Tidak Lolos Analisa Spasial",
                                   len(df_per_row[df_per_row["Status Analisa Spasial"] == "Tidak Lolos"])))
        write_grouped_sheet(ws_pr, df_per_row,
                            f"PER SUMUR: 1 baris = 1 nama sumur {per_row_label_p} "
                            f"(redundant dikelompokkan berurutan, warna biru)",
                            stats_a, stats_b, label_a, label_b, show_stats=False,
                            extra_stats=extra_stats_pr,
                            row_fills=per_row_fills(df_per_row))

    # Sheet 3: Overlap
    if len(df_overlap) > 0:
        ws_ov = wb.create_sheet("Overlap")
        extra_stats_overlap = [("Jumlah Titik Overlap", len(df_overlap))]
        if "Kesamaan Nama" in df_overlap.columns:
            extra_stats_overlap.append(("Memiliki Nama Sama", len(df_overlap[df_overlap["Kesamaan Nama"] == "Sama"])))
            extra_stats_overlap.append(("Memiliki Nama Berbeda", len(df_overlap[df_overlap["Kesamaan Nama"] == "Berbeda"])))
        if "Status Analisa Spasial" in df_overlap.columns:
            extra_stats_overlap.append(("Lolos Analisa Spasial", len(df_overlap[df_overlap["Status Analisa Spasial"] == "Lolos"])))
            extra_stats_overlap.append(("Tidak Lolos Analisa Spasial", len(df_overlap[df_overlap["Status Analisa Spasial"] == "Tidak Lolos"])))
        write_grouped_sheet(ws_ov, df_overlap,
                            f"OVERLAP: {label_a} × {label_b} (jarak ≤ {threshold_m} m)",
                            stats_a, stats_b, label_a, label_b, show_stats=True,
                            extra_stats=extra_stats_overlap)

    # Sheet 4: Hanya File A
    if len(df_file_a_only) > 0:
        ws_a = wb.create_sheet(f"Hanya {label_a}"[:31])
        write_grouped_sheet(ws_a, df_file_a_only, f"HANYA {label_a.upper()}",
                            stats_a, {}, label_a, label_b, show_stats=False)

    # Sheet 5: Hanya File B
    if len(df_file_b_only) > 0:
        ws_b = wb.create_sheet(f"Hanya {label_b}"[:31])
        write_grouped_sheet(ws_b, df_file_b_only, f"HANYA {label_b.upper()}",
                            {}, stats_b, label_a, label_b, show_stats=False)

    # Sheet 6: Lolos Analisa Spasial
    if df_lolos_spasial is not None and len(df_lolos_spasial) > 0:
        ws_lolos = wb.create_sheet("Lolos Analisa Spasial")
        extra_stats_lolos = [("Jumlah Titik Lolos Analisa Spasial", len(df_lolos_spasial))]
        if "Kesamaan Nama" in df_lolos_spasial.columns:
            extra_stats_lolos.append(("Memiliki Nama Sama", len(df_lolos_spasial[df_lolos_spasial["Kesamaan Nama"] == "Sama"])))
            extra_stats_lolos.append(("Memiliki Nama Berbeda", len(df_lolos_spasial[df_lolos_spasial["Kesamaan Nama"] == "Berbeda"])))
        write_grouped_sheet(ws_lolos, df_lolos_spasial, "LOLOS ANALISA SPASIAL",
                            stats_a, stats_b, label_a, label_b, show_stats=False,
                            extra_stats=extra_stats_lolos)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── UI ───────────────────────────────────────────────────────────────────────

st.title("🗺️ KML/KMZ Compare + Polygon Classifier")
st.caption("Overlap antar file = jarak haversine ≤ threshold. Koordinat A & B ditampilkan terpisah, full precision (string asli file). Duplikat intra-file default hanya koordinat identik persis. Sheet 'Per Sumur' = 1 baris per nama sumur file acuan (bisa dipilih), urutan asli.")

with st.sidebar:
    st.header("⚙️ Pengaturan")
    threshold_m = st.number_input(
        "Threshold overlap ANTAR file (meter)",
        min_value=0.1, max_value=10_000.0, value=5.0, step=1.0,
        help="Titik File 1 dan File 2 dianggap OVERLAP kalau jarak haversine ≤ nilai ini."
    )
    dup_threshold_m = st.number_input(
        "Threshold duplikat DALAM satu file (meter)",
        min_value=0.0, max_value=1_000.0, value=0.0, step=1.0,
        help="0 (default) = hanya titik dengan koordinat IDENTIK PERSIS yang digabung "
             "jadi satu lokasi. Sumur berbeda yang berdekatan (mis. 3-4 m) tetap "
             "jadi baris terpisah dengan koordinat masing-masing. "
             "Isi > 0 kalau memang mau gabungkan titik berdekatan dalam satu file."
    )
    st.caption(f"Overlap antar file: ≤ {threshold_m:g} m | Duplikat intra-file: "
               f"{'identik persis' if dup_threshold_m == 0 else f'≤ {dup_threshold_m:g} m'}")

    st.markdown("---")
    per_sumur_choice = st.radio(
        "Sheet 'Per Sumur' — baris ikut file mana?",
        ["Otomatis (file terbanyak)", "Ikut File 1", "Ikut File 2"],
        index=0,
        help="Acuan baris di sheet Per Sumur. Otomatis = file dengan jumlah nama sumur terbanyak. "
             "Pilih 'Ikut File 1' atau 'Ikut File 2' untuk memaksa baris mengikuti file tertentu.",
    )
    per_sumur_primary = {"Otomatis (file terbanyak)": "auto",
                         "Ikut File 1": "a", "Ikut File 2": "b"}[per_sumur_choice]

    st.markdown("---")
    use_polygon = st.checkbox("🗺️ Gunakan Polygon Classifier", value=False)

col1, col2 = st.columns(2)

with col1:
    st.subheader("📂 File 1")
    file_a = st.file_uploader("Upload KML/KMZ", type=["kml", "kmz"], key="file_a")
    label_a = filename_to_label(file_a) if file_a else "File 1"
    st.text_input("Label", value=label_a, key="label_a_display", disabled=True)

with col2:
    st.subheader("📂 File 2")
    file_b = st.file_uploader("Upload KML/KMZ", type=["kml", "kmz"], key="file_b")
    label_b = filename_to_label(file_b) if file_b else "File 2"
    st.text_input("Label", value=label_b, key="label_b_display", disabled=True)

# Polygon Classifier Section
polygon_slots = []
polygon_rules = {}

if use_polygon:
    st.markdown("---")
    st.subheader("🗺️ Polygon Classifier (1-5 Polygon)")

    rule_options = ["Hanya Info (tidak difilter)", "Lolos jika Dalam", "Lolos jika Luar"]

    for i in range(1, 6):
        required_tag = " *(wajib)*" if i == 1 else " *(opsional)*"
        with st.expander(f"Polygon {i}{required_tag}", expanded=(i == 1)):
            pfile = st.file_uploader(f"File Polygon {i}", type=['kml', 'kmz'],
                                      accept_multiple_files=False, key=f"poly_{i}",
                                      label_visibility="collapsed")
            if pfile:
                default_name = os.path.splitext(pfile.name)[0]
                pname = st.text_input(f"Nama Polygon {i}", value=default_name, key=f"pname_{i}")
                prule = st.selectbox(f"Rule Polygon {i}", rule_options, key=f"prule_{i}")
                polygon_slots.append({'idx': i, 'file': pfile, 'name': pname, 'rule': prule})

if st.button("🚀 PROSES", type="primary", use_container_width=True):

    if not file_a or not file_b:
        st.error("⚠️ Upload 2 file dulu.")
        st.stop()

    if use_polygon and not polygon_slots:
        st.error("⚠️ Polygon Classifier aktif, upload minimal 1 polygon.")
        st.stop()

    with st.spinner("Memproses..."):
        try:
            label_a = filename_to_label(file_a) or "File 1"
            label_b = filename_to_label(file_b) or "File 2"

            kml_bytes_a = extract_kml_bytes(file_a)
            kml_bytes_b = extract_kml_bytes(file_b)

            if kml_bytes_a is None or kml_bytes_b is None:
                st.stop()

            recs_a = parse_kml(kml_bytes_a, label_a)
            recs_b = parse_kml(kml_bytes_b, label_b)

            if not recs_a or not recs_b:
                st.error("❌ Tidak ada titik yang terbaca.")
                st.stop()

            # Info acuan Per Sumur yang dipakai
            if per_sumur_primary == "a":
                acuan_label = label_a
            elif per_sumur_primary == "b":
                acuan_label = label_b
            else:
                acuan_label = label_a if len(recs_a) >= len(recs_b) else label_b

            st.success(f"**{label_a}**: {len(recs_a)} titik | **{label_b}**: {len(recs_b)} titik | "
                       f"Overlap ≤ {threshold_m:g} m | Duplikat intra-file: "
                       f"{'identik persis' if dup_threshold_m == 0 else f'≤ {dup_threshold_m:g} m'} | "
                       f"Per Sumur ikut: **{acuan_label}**")

            (df_all, max_dup_a, max_dup_b, stats_a, stats_b, coord_cols,
             clusters_a, clusters_b, matches) = build_grouped_comparison(
                recs_a, recs_b, threshold_m, dup_threshold_m, label_a, label_b
            )

            df_overlap = df_all[df_all["Keterangan"] == "Overlap"].reset_index(drop=True)
            df_file_a_only = df_all[(df_all["Keterangan"] == "Tidak Overlap") &
                                    (df_all[f"Jumlah Sumur {label_a}"] > 0)].reset_index(drop=True)
            df_file_b_only = df_all[(df_all["Keterangan"] == "Tidak Overlap") &
                                    (df_all[f"Jumlah Sumur {label_b}"] > 0)].reset_index(drop=True)

            # Sheet Per Sumur: baris = tiap nama sumur di file ACUAN pilihan user, urutan asli
            df_per_row, per_row_label_p, per_row_label_s = build_per_row_comparison(
                recs_a, recs_b, clusters_a, clusters_b, matches, label_a, label_b,
                primary=per_sumur_primary
            )

            # Polygon Classifier
            polygon_cols_list = []
            df_lolos_spasial = None

            if use_polygon:
                for slot in polygon_slots:
                    poly_kml = extract_kml_bytes(slot['file'])
                    if poly_kml is None:
                        continue

                    polys = extract_polygons_from_kml(poly_kml)
                    if not polys:
                        st.warning(f"⚠️ Tidak ada polygon di file Polygon {slot['idx']}, dilewati.")
                        continue

                    union_poly = unary_union(polys)
                    col_name = f"Polygon {slot['idx']} ({slot['name']})"
                    polygon_cols_list.append(col_name)
                    polygon_rules[col_name] = slot['rule']

                    for df in [df_all, df_overlap, df_file_a_only, df_file_b_only, df_per_row]:
                        df[col_name] = df.apply(
                            lambda r: 'Dalam' if union_poly.covers(Point(r['_lon'], r['_lat'])) else 'Luar',
                            axis=1
                        )

                def check_lolos_spasial(row):
                    for col_name, rule in polygon_rules.items():
                        if col_name in row.index:
                            status = row[col_name]
                            if rule == "Lolos jika Dalam" and status != 'Dalam':
                                return "Tidak Lolos"
                            elif rule == "Lolos jika Luar" and status != 'Luar':
                                return "Tidak Lolos"
                    return "Lolos"

                for df in [df_all, df_overlap, df_file_a_only, df_file_b_only, df_per_row]:
                    df["Status Analisa Spasial"] = df.apply(check_lolos_spasial, axis=1)

                df_lolos_spasial = df_overlap[df_overlap["Status Analisa Spasial"] == "Lolos"].reset_index(drop=True)

            # Metrics
            m1, m2, m3, m4 = st.columns(4)
            m1.metric(f"Lokasi {label_a}", stats_a["total_koordinat"])
            m2.metric(f"Lokasi {label_b}", stats_b["total_koordinat"])
            m3.metric("Overlap", len(df_overlap))
            if use_polygon and df_lolos_spasial is not None:
                m4.metric("Lolos Spasial", len(df_lolos_spasial))
            else:
                m4.metric("Max Duplikat", f"{max_dup_a}/{max_dup_b}")

            st.markdown("---")

            def preview(df):
                return df[[c for c in df.columns if not c.startswith("_")]]

            tab_labels = [
                f"Semua ({len(df_all)})",
                f"Per Sumur ({len(df_per_row)})",
                f"Overlap ({len(df_overlap)})",
                f"Hanya {label_a} ({len(df_file_a_only)})",
                f"Hanya {label_b} ({len(df_file_b_only)})",
            ]
            if use_polygon:
                tab_labels.append(f"Lolos Spasial ({len(df_lolos_spasial)})")

            tabs = st.tabs(tab_labels)

            with tabs[0]:
                st.dataframe(preview(df_all), use_container_width=True, height=400)
            with tabs[1]:
                st.caption(f"1 baris = 1 nama sumur **{per_row_label_p}** (file acuan). "
                           f"Redundant (koordinat sama) dikelompokkan berurutan → di Excel diberi warna biru. "
                           f"Nama {per_row_label_s} digabung pakai koma.")
                st.dataframe(preview(df_per_row), use_container_width=True, height=400)
            with tabs[2]:
                st.dataframe(preview(df_overlap), use_container_width=True, height=400)
            with tabs[3]:
                st.dataframe(preview(df_file_a_only), use_container_width=True, height=400) if len(df_file_a_only) > 0 else st.info("Kosong")
            with tabs[4]:
                st.dataframe(preview(df_file_b_only), use_container_width=True, height=400) if len(df_file_b_only) > 0 else st.info("Kosong")

            if use_polygon:
                with tabs[5]:
                    st.dataframe(preview(df_lolos_spasial), use_container_width=True, height=400) if len(df_lolos_spasial) > 0 else st.info("Kosong")

            st.markdown("---")

            excel_bytes = build_excel_grouped(
                df_all, df_overlap, df_file_a_only, df_file_b_only, df_lolos_spasial,
                df_per_row, per_row_label_p,
                label_a, label_b, stats_a, stats_b, threshold_m, dup_threshold_m,
                coord_cols, polygon_cols_list
            )

            fn = f"Compare_{label_a}_vs_{label_b}.xlsx".replace(" ", "_")
            st.download_button(
                label="📥 Download Excel",
                data=excel_bytes,
                file_name=fn,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
            )

        except Exception as e:
            st.error(f"❌ Error: {e}")
            import traceback
            st.code(traceback.format_exc())
